import contextlib
import datetime as dt
import hashlib
import io
import json
from pathlib import Path
import sqlite3
import subprocess
import sys
import tempfile
import time
import unittest
from unittest import mock

from continue_catalog_sync_loader import sync


def exact_record(candidate_id="b" * 28):
    return {
        "candidate_id": candidate_id,
        "status": "exact",
        "reasons": [],
        "audio_match_kind": "exact_title",
        "audio_match_score": 100,
        "spotify_id": "1" * 22,
        "spotify_duration_seconds": 120.0,
        "spotify_match_kind": "exact",
        "track": {
            "source_row": 1,
            "release": "Release",
            "upc": "12345678",
            "title": "Track",
            "artists": ["Artist"],
            "duration_seconds": 120.0,
            "isrc": "USABC1200001",
            "genre": "Lofi",
            "subgenre": "",
        },
        "audio": {
            "file_id": "1" + "D" * 24,
            "name": "Track.wav",
            "folder_id": "",
            "folder_path": "",
            "mime_type": "audio/wav",
            "size_bytes": 100,
            "modified_time": "",
        },
        "cover": {"file_id": "1" + "C" * 24, "is_square": True, "quality": "3000 x 3000", "upc": "12345678"},
        "inspection": {
            "status": "complete",
            "mode": "full",
            "content_type": "audio/wav",
            "content_length": 100,
            "sha256": "a" * 64,
            "error": None,
            "wav": {
                "container": "RIFF",
                "codec": "PCM",
                "channels": 2,
                "sample_rate": 48000,
                "bits_per_sample": 16,
                "byte_rate": 192000,
                "data_size": 23040000,
                "duration_seconds": 120.0,
            },
        },
    }


def spotify_evidence_record(
    candidate_id="b" * 28,
    *,
    disposition="accepted",
    title="Track",
):
    accepted = disposition == "accepted"
    return {
        "recordKey": candidate_id,
        "inputIndex": 0,
        "spotifyId": "1" * 22,
        "local": {
            "title": "Track",
            "artists": ["Artist"],
            "releaseTitle": "Release",
            "upc": "12345678",
            "sourceSha256": "a" * 64,
            "audioInspectionComplete": True,
        },
        "spotify": {
            "title": title,
            "artists": ["Artist"],
            "durationMs": 120000,
            "sources": {"oembed": "ok", "embed": "ok"},
        },
        "cache": "miss",
        "disposition": disposition,
        "reasons": [] if accepted else ["manual_review"],
        "checks": {
            "title": {"status": "exact" if accepted else "review"},
            "artists": {"status": "exact" if accepted else "review"},
            "duration": {"status": "match" if accepted else "review"},
        },
        "artworkRecommendation": "owned",
    }


def owner_attestation_payload(
    drive_root="ABCDEFGHIJKLMNO",
    workbook_id="S" * 20,
    **overrides,
):
    payload = {
        "schemaVersion": 1,
        "kind": "catalog_owner_attestation",
        "approved": True,
        "claims": {
            "catalogueAlreadyReleased": True,
            "rightsToPublishFullLengthListeningCopies": True,
            "rightsToOfferLicensedDownloads": True,
            "humanMadeNoGenerativeAI": True,
        },
        "scope": {
            "catalogueSourceUrl": "https://lofi-records.netlify.app/#/catalog",
            "workbookSourceUrl": f"https://docs.google.com/spreadsheets/d/{workbook_id}/edit",
            "driveSourceFolderId": drive_root,
        },
        "reviewer": "Catalogue owner",
        "reviewerRole": "catalogue_owner",
        "reviewedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
    }
    payload.update(overrides)
    return payload


class InventoryTests(unittest.TestCase):
    def test_public_inventory_summary_never_returns_rows_or_ids(self):
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "inventory.json"
            private_id = "A" * 20
            destination.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "complete": False,
                        "files": [
                            {"id": private_id, "name": "Private track.wav", "mimeType": "audio/wav"},
                            {"id": "B" * 20, "name": "Private folder", "mimeType": "application/vnd.google-apps.folder"},
                        ],
                    }
                ),
                encoding="utf-8",
            )
            summary = sync.drive_inventory_summary(destination)
            self.assertEqual(summary["wavFiles"], 1)
            self.assertEqual(summary["items"], 2)
            self.assertFalse(summary["complete"])
            serialized_summary = json.dumps(summary)
            self.assertNotIn(private_id, serialized_summary)
            self.assertNotIn("Private track", serialized_summary)
    def test_drive_sync_command_uses_private_seed_and_resumable_public_scraper(self):
        with tempfile.TemporaryDirectory() as directory:
            temporary = Path(directory)
            (temporary / "sync_lofi_drive.py").write_text("# fixture\n", encoding="utf-8")
            original = sync.AUDIT_DIRECTORY
            sync.AUDIT_DIRECTORY = temporary
            try:
                command = sync.build_drive_sync_command(
                    {
                        "drive_seed": temporary / "seed.json",
                        "inventory_directory": temporary / "output",
                        "inventory_release_batch": 25,
                    }
                )
            finally:
                sync.AUDIT_DIRECTORY = original
        self.assertIn("--resume", command)
        self.assertIn("--max-releases", command)
        self.assertNotIn("--token", command)
        self.assertNotIn("--api-key", command)

    def test_global_inventory_drain_runs_sequential_checkpoints_until_exhausted(self):
        documents = [
            [{"status": "partial", "mode": "apply", "pendingReleases": 150, "inventoryItems": 100, "errors": 0, "complete": False}],
            [{"status": "partial", "mode": "apply", "pendingReleases": 50, "inventoryItems": 200, "errors": 0, "complete": False}],
            [{"status": "ok", "mode": "apply", "pendingReleases": 0, "inventoryItems": 250, "errors": 0, "complete": True}],
        ]
        config = {
            "drive_seed": Path("seed.json"),
            "inventory_directory": Path("inventory"),
            "inventory_release_batch": 50,
        }
        with mock.patch.object(sync, "run_step", side_effect=documents) as runner:
            summary = sync.drain_drive_inventory(config)
        self.assertEqual(runner.call_count, 3)
        self.assertEqual(summary["passes"], 3)
        self.assertTrue(summary["complete"])
        for call in runner.call_args_list:
            command = call.args[1]
            self.assertEqual(
                command[command.index("--max-releases") + 1],
                str(sync.DIRECT_INVENTORY_RELEASE_BATCH),
            )

    def test_drive_summary_keeps_real_aggregate_fields_and_drops_private_rows(self):
        summary = sync.compact_child_summary(
            "drive_sync",
            [
                {
                    "schemaVersion": 1,
                    "mode": "apply",
                    "status": "partial",
                    "seedReleases": 743,
                    "rootFolders": 744,
                    "pendingReleases": 719,
                    "pendingItems": 12,
                    "inventoryItems": 90,
                    "audioFiles": 60,
                    "artworkFiles": 30,
                    "complete": False,
                    "rootNew": 1,
                    "selected": 25,
                    "scanned": 24,
                    "errors": 1,
                    "privateTitle": "Never expose this",
                    "folderId": "A" * 20,
                }
            ],
        )
        self.assertEqual(
            summary,
            {
                "seedReleases": 743,
                "rootFolders": 744,
                "pendingReleases": 719,
                "pendingItems": 12,
                "inventoryItems": 90,
                "audioFiles": 60,
                "artworkFiles": 30,
                "rootNew": 1,
                "selected": 25,
                "scanned": 24,
                "errors": 1,
                "complete": False,
                "mode": "apply",
                "status": "partial",
            },
        )
        self.assertNotIn("Never expose", json.dumps(summary))
        self.assertNotIn("A" * 20, json.dumps(summary))

    def test_node_resolution_finds_sibling_codex_runtime_when_path_is_empty(self):
        with tempfile.TemporaryDirectory() as directory:
            dependencies = Path(directory) / "dependencies"
            python = dependencies / "python" / "python.exe"
            node = dependencies / "node" / "bin" / "node.exe"
            python.parent.mkdir(parents=True)
            node.parent.mkdir(parents=True)
            python.write_bytes(b"")
            node.write_bytes(b"")
            resolved = sync.resolve_node_executable(
                environment={},
                which=lambda _name: None,
                python_executable=python,
            )
        self.assertEqual(resolved, str(node.resolve()))

    def test_node_resolution_honors_valid_override_and_rejects_injection(self):
        with tempfile.TemporaryDirectory() as directory:
            node = Path(directory) / "node.exe"
            node.write_bytes(b"")
            self.assertEqual(
                sync.resolve_node_executable(
                    environment={"SYMBIOME_NODE_EXECUTABLE": str(node)},
                    which=lambda _name: None,
                ),
                str(node.resolve()),
            )
        with self.assertRaisesRegex(sync.SyncError, "node_executable_invalid"):
            sync.resolve_node_executable(
                environment={"SYMBIOME_NODE_EXECUTABLE": "node\nmalicious"},
                which=lambda _name: None,
            )

    def test_google_sheet_refresh_is_host_allowlisted_and_xlsx_validated(self):
        import openpyxl

        identifier = "S" * 20
        export = sync.google_sheet_export_url(
            f"https://docs.google.com/spreadsheets/d/{identifier}/edit?usp=sharing"
        )
        self.assertEqual(
            export,
            f"https://docs.google.com/spreadsheets/d/{identifier}/export?format=xlsx",
        )
        with self.assertRaisesRegex(sync.SyncError, "workbook_source_url_invalid"):
            sync.google_sheet_export_url("https://example.com/private.xlsx")

        workbook = openpyxl.Workbook()
        workbook.active.title = "Publishing catalogue"
        workbook.create_sheet("Cover Album")
        workbook.create_sheet("Identity Info")
        payload = io.BytesIO()
        workbook.save(payload)
        workbook.close()

        class Response(io.BytesIO):
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                self.close()

        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "all-data.xlsx"
            result = sync.refresh_workbook(
                f"https://docs.google.com/spreadsheets/d/{identifier}/edit",
                destination,
                opener=lambda *_args, **_kwargs: Response(payload.getvalue()),
            )
            self.assertTrue(result["changed"])
            self.assertTrue(destination.is_file())


class SingletonLockTests(unittest.TestCase):
    def test_concurrent_run_fails_fast_without_overwriting_state_and_crash_releases_lock(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            work_directory = root / "private-work"
            ready = root / "ready"
            config_path = root / "config.json"
            last_run = work_directory / "last-run.json"
            work_directory.mkdir()
            sentinel = '{"status":"previous"}\n'
            last_run.write_text(sentinel, encoding="utf-8")

            child_code = (
                "import sys,time\n"
                "from pathlib import Path\n"
                "sys.path.insert(0, sys.argv[1])\n"
                "import continue_catalog_sync as sync\n"
                "with sync.exclusive_sync_lock(Path(sys.argv[2])):\n"
                " Path(sys.argv[3]).write_text('ready', encoding='utf-8')\n"
                " time.sleep(60)\n"
            )
            child = subprocess.Popen(
                [
                    sys.executable,
                    "-c",
                    child_code,
                    str(sync.AUDIT_DIRECTORY),
                    str(work_directory),
                    str(ready),
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            try:
                deadline = time.monotonic() + 10
                while not ready.is_file() and child.poll() is None and time.monotonic() < deadline:
                    time.sleep(0.02)
                self.assertTrue(ready.is_file(), f"lock holder exited with {child.poll()}")

                output = io.StringIO()
                config = {"work_directory": work_directory}
                with mock.patch.object(sync, "apply_low_resource_priority") as priority:
                    with mock.patch.object(sync, "load_config", return_value=config):
                        with mock.patch.object(sync, "execute") as execute:
                            with contextlib.redirect_stdout(output):
                                status = sync.main(
                                    ["--config", str(config_path), "--mode", "continue", "--allow-network"]
                                )

                self.assertEqual(status, 1)
                priority.assert_called_once_with()
                self.assertFalse(execute.called)
                self.assertEqual(json.loads(output.getvalue())["error"], "sync_already_running")
                self.assertEqual(last_run.read_text(encoding="utf-8"), sentinel)

                child.kill()
                child.wait(timeout=10)
                with sync.exclusive_sync_lock(work_directory):
                    pass
                self.assertEqual((work_directory / sync.SYNC_LOCK_FILENAME).read_bytes(), b"\0")
            finally:
                if child.poll() is None:
                    child.kill()
                    child.wait(timeout=10)


class LowResourceTests(unittest.TestCase):
    def test_priority_is_lowered_on_windows_and_posix_without_propagating_failures(self):
        process = object()
        kernel32 = mock.Mock()
        kernel32.GetCurrentProcess.return_value = process
        kernel32.SetPriorityClass.return_value = 1
        factory = mock.Mock(return_value=kernel32)

        self.assertTrue(
            sync.apply_low_resource_priority(
                platform_name="nt",
                windows_api_factory=factory,
            )
        )
        factory.assert_called_once_with("kernel32", use_last_error=True)
        kernel32.SetPriorityClass.assert_called_once_with(
            process,
            sync.BELOW_NORMAL_PRIORITY_CLASS,
        )

        nice = mock.Mock()
        self.assertTrue(
            sync.apply_low_resource_priority(
                platform_name="posix",
                nice_function=nice,
            )
        )
        nice.assert_called_once_with(sync.LOW_RESOURCE_NICE_INCREMENT)

        failing_nice = mock.Mock(side_effect=OSError("fixture"))
        self.assertFalse(
            sync.apply_low_resource_priority(
                platform_name="posix",
                nice_function=failing_nice,
            )
        )

    def test_thread_caps_preserve_one_and_replace_larger_or_invalid_values(self):
        source = {
            "PATH": "fixture-path",
            "OMP_NUM_THREADS": "1",
            "MKL_NUM_THREADS": "8",
            "OPENBLAS_NUM_THREADS": "invalid",
            "NUMEXPR_NUM_THREADS": "4",
            "VECLIB_MAXIMUM_THREADS": "2",
        }
        limited = sync.low_resource_child_environment(source)

        self.assertEqual(limited["PATH"], "fixture-path")
        for key in sync.LOW_RESOURCE_THREAD_ENVIRONMENT_KEYS:
            self.assertEqual(limited[key], "1")
        self.assertEqual(source["MKL_NUM_THREADS"], "8")

    def test_python_node_and_ffmpeg_children_inherit_single_thread_caps(self):
        completed = sync.subprocess.CompletedProcess(
            args=["fixture"], returncode=0, stdout="{}", stderr=""
        )
        commands = (
            [sys.executable, "child.py"],
            ["node", "child.mjs"],
            ["ffmpeg", "-version"],
        )
        for command in commands:
            with self.subTest(command=command[0]):
                with mock.patch.object(sync.subprocess, "run", return_value=completed) as runner:
                    sync.run_step(
                        "fixture",
                        command,
                        environment={"PATH": "fixture-path", "OMP_NUM_THREADS": "16"},
                    )
                child_environment = runner.call_args.kwargs["env"]
                for key in sync.LOW_RESOURCE_THREAD_ENVIRONMENT_KEYS:
                    self.assertEqual(child_environment[key], "1")


class PublicationCredentialTests(unittest.TestCase):
    def test_secrets_are_removed_from_every_default_child_environment(self):
        completed = sync.subprocess.CompletedProcess(
            args=["fixture"], returncode=0, stdout="{}", stderr=""
        )
        parent_environment = {
            "PATH": "fixture-path",
            "CATALOG_PIPELINE_TOKEN": "pipeline-secret-fixture-123456",
            "OAI_SITES_AUTHORIZATION": "sites-secret-fixture-123456",
        }
        with mock.patch.dict(sync.os.environ, parent_environment, clear=True):
            with mock.patch.object(sync.subprocess, "run", return_value=completed) as runner:
                sync.run_step("fixture", ["fixture-command"])
        child_environment = runner.call_args.kwargs["env"]
        self.assertEqual(child_environment["PATH"], "fixture-path")
        self.assertNotIn("CATALOG_PIPELINE_TOKEN", child_environment)
        self.assertNotIn("OAI_SITES_AUTHORIZATION", child_environment)

    def test_publication_credentials_are_injected_only_via_apply_environment(self):
        pipeline_token = "pipeline-secret-fixture-123456"
        sites_authorization = "sites-secret-fixture-123456"
        environment = sync.publication_apply_environment(
            {
                "pipeline_token": pipeline_token,
                "sites_authorization": sites_authorization,
            },
            environment={"PATH": "fixture-path"},
        )
        command = ["process_catalog.py", "--apply"]
        completed = sync.subprocess.CompletedProcess(
            args=command, returncode=0, stdout="{}", stderr=""
        )
        with mock.patch.object(sync.subprocess, "run", return_value=completed) as runner:
            sync.run_step("publication_apply", command, environment=environment)
        child_environment = runner.call_args.kwargs["env"]
        self.assertEqual(child_environment["CATALOG_PIPELINE_TOKEN"], pipeline_token)
        self.assertEqual(child_environment["OAI_SITES_AUTHORIZATION"], sites_authorization)
        self.assertNotIn(pipeline_token, " ".join(command))
        self.assertNotIn(sites_authorization, " ".join(command))

    def test_publication_credentials_never_escape_through_child_output(self):
        pipeline_token = "pipeline-secret-fixture-123456"
        sites_authorization = "sites-secret-fixture-123456"
        environment = sync.publication_apply_environment(
            {
                "pipeline_token": pipeline_token,
                "sites_authorization": sites_authorization,
            },
            environment={},
        )
        completed = sync.subprocess.CompletedProcess(
            args=["fixture"],
            returncode=0,
            stdout=json.dumps({"accidental": pipeline_token}),
            stderr="",
        )
        with mock.patch.object(sync.subprocess, "run", return_value=completed):
            with self.assertRaisesRegex(sync.SyncError, "publication_apply_sensitive_output_blocked") as caught:
                sync.run_step("publication_apply", ["fixture"], environment=environment)
        self.assertNotIn(pipeline_token, str(caught.exception))
        self.assertNotIn(sites_authorization, str(caught.exception))

    def test_publication_credentials_fail_closed_when_incomplete(self):
        with self.assertRaisesRegex(sync.SyncError, "publication_credentials_missing"):
            sync.publication_apply_environment(
                {"pipeline_token": "pipeline-secret-fixture-123456", "sites_authorization": None},
                environment={},
            )


class PublicationPartialTests(unittest.TestCase):
    @staticmethod
    def partial_document():
        return {
            "mode": "apply",
            "step": "process",
            "counts": {
                "selected": 5,
                "published": 2,
                "already_published": 1,
                "promotion_blocked": 1,
                "failed": 1,
            },
            "pipelineState": {"published": 3, "failed": 1, "promotion_blocked": 1},
        }

    def test_return_code_two_is_accepted_only_for_publication_apply(self):
        completed = sync.subprocess.CompletedProcess(
            args=["fixture"],
            returncode=2,
            stdout=json.dumps(self.partial_document()),
            stderr="",
        )
        with mock.patch.object(sync.subprocess, "run", return_value=completed):
            with self.assertRaisesRegex(sync.SyncError, "publication_apply_failed"):
                sync.run_step("publication_apply", ["fixture"])
        with mock.patch.object(sync.subprocess, "run", return_value=completed):
            documents, return_code = sync.run_step_result(
                "publication_apply",
                ["fixture"],
                accepted_return_codes=(0, 2),
            )
        self.assertEqual(return_code, 2)
        self.assertEqual(documents[-1]["counts"]["failed"], 1)
        with self.assertRaisesRegex(sync.SyncError, "subprocess_return_code_policy_invalid"):
            sync.run_step_result("spotify", ["fixture"], accepted_return_codes=(0, 2))

    def test_partial_summary_preserves_aggregate_counts_and_is_resumable_success(self):
        summary, partial = sync.publication_apply_summary([self.partial_document()], 2)
        self.assertTrue(partial)
        self.assertEqual(summary["counts"]["published"], 2)
        self.assertEqual(summary["counts"]["promotion_blocked"], 1)
        self.assertEqual(summary["counts"]["failed"], 1)
        self.assertEqual(summary["pipelineState"]["published"], 3)

        work_directory = sync.AUDIT_DIRECTORY / "private" / "partial-status-fixture"
        config_path = work_directory / "config.json"
        config = {"work_directory": work_directory}
        output = io.StringIO()
        with mock.patch.object(sync, "load_config", return_value=config):
            with mock.patch.object(sync, "exclusive_sync_lock", return_value=contextlib.nullcontext()):
                with mock.patch.object(sync, "atomic_write_json"):
                    with mock.patch.object(
                        sync,
                        "execute",
                        return_value={"schemaVersion": 1, "status": "partial", "resumable": True},
                    ):
                        with contextlib.redirect_stdout(output):
                            exit_code = sync.main(["--config", str(config_path), "--mode", "publish"])
        self.assertEqual(exit_code, 0)
        self.assertEqual(json.loads(output.getvalue())["status"], "partial")

    def test_return_code_and_summary_must_agree(self):
        complete = self.partial_document()
        complete["counts"] = {
            "selected": 2,
            "published": 2,
            "already_published": 0,
            "promotion_blocked": 0,
            "failed": 0,
        }
        with self.assertRaisesRegex(sync.SyncError, "publication_apply_summary_invalid"):
            sync.publication_apply_summary([complete], 2)
        with self.assertRaisesRegex(sync.SyncError, "publication_apply_summary_invalid"):
            sync.publication_apply_summary([self.partial_document()], 0)


class ThroughputBoundsTests(unittest.TestCase):
    def test_accelerated_hourly_bounds_accept_safe_targets_and_reject_larger_batches(self):
        targets = (
            ("driveInventoryReleasesPerRun", 50, sync.MAX_INVENTORY_RELEASE_BATCH),
            ("inspectionBatchSize", 50, sync.MAX_INSPECTION_BATCH),
            ("spotifyBatchSize", 50, sync.MAX_SPOTIFY_BATCH),
            ("publicationBatchSize", 25, sync.MAX_PUBLICATION_BATCH),
            ("maximumReleasesPerRun", 10, sync.MAX_RELEASE_BATCH),
            ("maximumRunMinutes", 55, sync.MAX_RUN_MINUTES),
        )
        for key, target, maximum in targets:
            with self.subTest(key=key):
                self.assertEqual(sync.bounded_integer({key: target}, key, 1, maximum), target)
                with self.assertRaisesRegex(sync.SyncError, f"{key.casefold()}_out_of_bounds"):
                    sync.bounded_integer({key: maximum + 1}, key, 1, maximum)

    def test_accelerated_inspection_remains_one_sequential_bounded_invocation(self):
        config = {
            "workbook": Path("private.xlsx"),
            "orchard": None,
            "work_directory": Path("private"),
            "inspection_batch": 50,
            "release_batch": 10,
            "maximum_run_minutes": 55,
        }
        command = sync.build_ingest_command(
            config, Path("inventory.json"), apply=True, inspect_full=True
        )
        self.assertEqual(command[command.index("--batch-size") + 1], "50")
        self.assertEqual(command[command.index("--release-batch-size") + 1], "10")
        self.assertEqual(command[command.index("--max-inspection-seconds") + 1], "3300")
        self.assertEqual(command.count("--allow-network"), 1)

    def test_ingest_command_adds_private_central_snapshot_when_available(self):
        central = Path("private") / "central.json"
        baseline = Path("private") / "central-baseline.json"
        config = {
            "workbook": Path("private.xlsx"),
            "orchard": None,
            "work_directory": Path("private"),
            "central_inventory": central,
            "central_baseline_inventory": baseline,
            "inspection_batch": 1,
            "release_batch": 1,
            "maximum_run_minutes": 1,
        }
        command = sync.build_ingest_command(
            config, Path("inventory.json"), apply=False
        )
        self.assertEqual(
            command[command.index("--central-drive-inventory") + 1], str(central)
        )
        self.assertEqual(
            command[command.index("--central-drive-baseline-inventory") + 1],
            str(baseline),
        )


class FfmpegTests(unittest.TestCase):
    def test_private_config_accepts_optional_existing_ffmpeg_and_rejects_missing_file(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "all-data.xlsx"
            seed = root / "seed.json"
            ffmpeg = root / "ffmpeg.exe"
            config_path = root / "config.json"
            workbook.write_bytes(b"fixture")
            seed.write_text("{}", encoding="utf-8")
            ffmpeg.write_bytes(b"fixture")
            payload = {
                "schemaVersion": 1,
                "workDirectory": str(root / "work"),
                "workbook": str(workbook),
                "workbookSourceUrl": "https://docs.google.com/spreadsheets/d/" + "S" * 20,
                "driveSeed": str(seed),
                "driveInventoryDirectory": str(root / "inventory"),
                "ffmpegExecutable": str(ffmpeg),
                "pipelineToken": "pipeline-token-fixture-123456",
                "sitesAuthorization": "sites-authorization-fixture-123456",
            }
            config_path.write_text(json.dumps(payload), encoding="utf-8")
            config = sync.load_config(config_path)
            self.assertEqual(config["ffmpeg_executable"], ffmpeg.resolve())
            self.assertEqual(config["pipeline_token"], payload["pipelineToken"])
            self.assertEqual(config["sites_authorization"], payload["sitesAuthorization"])

            payload["pipelineToken"] = "invalid\nsecret-value"
            config_path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(sync.SyncError, "pipeline_token_invalid"):
                sync.load_config(config_path)
            payload["pipelineToken"] = "pipeline-token-fixture-123456"

            payload["ffmpegExecutable"] = str(root / "missing-ffmpeg.exe")
            config_path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(sync.SyncError, "ffmpeg_executable_missing"):
                sync.load_config(config_path)

    def test_ffmpeg_resolution_supports_override_imageio_and_path_then_fails_closed(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            override = root / "override.exe"
            imageio = root / "imageio.exe"
            path_ffmpeg = root / "path.exe"
            for executable in (override, imageio, path_ffmpeg):
                executable.write_bytes(b"fixture")

            self.assertEqual(
                sync.resolve_ffmpeg_executable(
                    None,
                    environment={"SYMBIOME_FFMPEG_EXECUTABLE": str(override)},
                    which=lambda _name: None,
                    imageio_getter=lambda: None,
                ),
                override.resolve(),
            )
            self.assertEqual(
                sync.resolve_ffmpeg_executable(
                    None,
                    environment={},
                    which=lambda _name: None,
                    imageio_getter=lambda: str(imageio),
                ),
                imageio.resolve(),
            )
            self.assertEqual(
                sync.resolve_ffmpeg_executable(
                    None,
                    environment={},
                    which=lambda name: str(path_ffmpeg) if name == "ffmpeg" else None,
                    imageio_getter=lambda: None,
                ),
                path_ffmpeg.resolve(),
            )
            with self.assertRaisesRegex(sync.SyncError, "ffmpeg_missing"):
                sync.resolve_ffmpeg_executable(
                    None,
                    environment={},
                    which=lambda _name: None,
                    imageio_getter=lambda: None,
                )
            with self.assertRaisesRegex(sync.SyncError, "ffmpeg_executable_invalid"):
                sync.resolve_ffmpeg_executable(
                    None,
                    environment={"SYMBIOME_FFMPEG_EXECUTABLE": "ffmpeg\nmalicious"},
                    which=lambda _name: None,
                    imageio_getter=lambda: None,
                )

    def test_process_commands_always_receive_ffmpeg_without_leaking_it_to_summary(self):
        private_ffmpeg = Path("private") / "ffmpeg.exe"
        common = {
            "exact_manifest": Path("private") / "selection.jsonl",
            "pipeline_state": Path("private") / "pipeline.sqlite3",
            "spotify_enrichment": Path("private") / "spotify.json",
            "ffmpeg_executable": private_ffmpeg,
        }
        dry_run = sync.build_process_catalog_command(**common, apply=False)
        apply = sync.build_process_catalog_command(**common, apply=True)
        for command in (dry_run, apply):
            self.assertIn("--ffmpeg", command)
            self.assertEqual(command[command.index("--ffmpeg") + 1], str(private_ffmpeg))
        self.assertIn("--dry-run", dry_run)
        self.assertNotIn("--rights-cleared", dry_run)
        self.assertIn("--apply", apply)
        self.assertIn("--rights-cleared", apply)
        summary = sync.compact_child_summary(
            "process",
            [{"mode": "dry-run", "selected": 3, "ffmpeg": str(private_ffmpeg)}],
        )
        self.assertNotIn("ffmpeg", summary)
        self.assertNotIn(str(private_ffmpeg), json.dumps(summary))

    def test_owner_direct_process_command_has_sealed_evidence_and_no_spotify_gate(self):
        command = sync.build_process_catalog_command(
            exact_manifest=Path("private") / "direct.jsonl",
            pipeline_state=Path("private") / "pipeline.sqlite3",
            spotify_enrichment=None,
            ffmpeg_executable=Path("private") / "ffmpeg.exe",
            verification_mode="catalog_owner_direct",
            owner_attestation_sha256="1" * 64,
            catalogue_scope_sha256="2" * 64,
            selection_sha256="3" * 64,
            apply=True,
        )
        self.assertNotIn("--spotify-enrichment", command)
        self.assertEqual(command[command.index("--verification-mode") + 1], "catalog_owner_direct")
        self.assertEqual(command[command.index("--batch-key") + 1], sync.DIRECT_BATCH_KEY)
        self.assertIn("--rights-cleared", command)
        self.assertIn("--human-made-cleared", command)


class SelectionTests(unittest.TestCase):
    def test_spotify_selection_requires_full_hash_and_rotates_by_attempt_count(self):
        first = exact_record("b" * 28)
        second = exact_record("c" * 28)
        first["spotify_duration_seconds"] = None
        second["spotify_duration_seconds"] = None
        state = {"spotifyAttempts": {first["candidate_id"]: {"count": 2, "lastAttemptAt": "now"}}}
        selected = sync.select_spotify_batch([first, second], state, 1)
        self.assertEqual(selected[0]["candidate_id"], second["candidate_id"])
        second["inspection"]["mode"] = "range"
        selected = sync.select_spotify_batch([first, second], state, 2)
        self.assertEqual([item["candidate_id"] for item in selected], [first["candidate_id"]])

    def test_owner_direct_selection_skips_identical_historical_published_base(self):
        published = exact_record("b" * 28)
        fresh = exact_record("c" * 28)
        with tempfile.TemporaryDirectory() as directory:
            state = Path(directory) / "pipeline.sqlite3"
            connection = sync.process.open_pipeline_state(state)
            connection.execute(
                """INSERT INTO pipeline_items (
                candidate_id, manifest_fingerprint, batch_key, status,
                rights_cleared_ack, human_made_cleared_ack, created_at, updated_at
                ) VALUES (?, ?, 'historical', 'published', 1, 1, ?, ?)""",
                (published["candidate_id"], "f" * 64, sync.utc_now(), sync.utc_now()),
            )
            connection.execute(
                """CREATE TABLE published_manifest_bases (
                candidate_id TEXT PRIMARY KEY,
                base_fingerprint TEXT NOT NULL,
                source_manifest_fingerprint TEXT NOT NULL,
                verified_at TEXT NOT NULL
                )"""
            )
            connection.execute(
                "INSERT INTO published_manifest_bases VALUES (?, ?, ?, ?)",
                (
                    published["candidate_id"],
                    sync.process.canonical_fingerprint(published),
                    "f" * 64,
                    sync.utc_now(),
                ),
            )
            connection.commit()
            connection.close()
            selected = sync.select_unpublished_direct_records(
                [published, fresh],
                state,
            )
        self.assertEqual([record["candidate_id"] for record in selected], [fresh["candidate_id"]])

    def test_publication_skips_only_identical_completed_fingerprint(self):
        record = exact_record()
        with tempfile.TemporaryDirectory() as directory:
            state_path = Path(directory) / "pipeline.sqlite3"
            connection = sqlite3.connect(state_path)
            connection.execute("CREATE TABLE pipeline_items (candidate_id TEXT, manifest_fingerprint TEXT, status TEXT)")
            connection.execute(
                "INSERT INTO pipeline_items VALUES (?, ?, 'published')",
                (record["candidate_id"], sync.process.canonical_fingerprint(record)),
            )
            connection.commit()
            connection.close()
            self.assertEqual(sync.select_publication_batch([record], state_path, 1), [])
            changed = json.loads(json.dumps(record))
            changed["track"]["genre"] = "Ambient"
            self.assertEqual(len(sync.select_publication_batch([changed], state_path, 1)), 1)

    def test_publication_prioritizes_fresh_exacts_then_rotates_failures(self):
        failed_many = exact_record("b" * 28)
        fresh_first = exact_record("c" * 28)
        failed_newer = exact_record("d" * 28)
        fresh_second = exact_record("e" * 28)
        failed_older = exact_record("f" * 28)
        records = [failed_many, fresh_first, failed_newer, fresh_second, failed_older]
        with tempfile.TemporaryDirectory() as directory:
            state_path = Path(directory) / "pipeline.sqlite3"
            connection = sync.process.open_pipeline_state(state_path)
            rows = (
                (failed_many, 4, "2026-01-01T00:00:00+00:00"),
                (failed_newer, 1, "2026-03-01T00:00:00+00:00"),
                (failed_older, 1, "2026-02-01T00:00:00+00:00"),
            )
            for record, attempts, updated_at in rows:
                connection.execute(
                    """INSERT INTO pipeline_items (
                    candidate_id, manifest_fingerprint, batch_key, status,
                    attempts, created_at, updated_at
                    ) VALUES (?, ?, 'old-batch', 'failed', ?, ?, ?)""",
                    (
                        record["candidate_id"],
                        sync.process.canonical_fingerprint(record),
                        attempts,
                        updated_at,
                        updated_at,
                    ),
                )
            connection.commit()
            connection.close()

            first_batch = sync.select_publication_batch(records, state_path, 2)
            self.assertEqual(
                [record["candidate_id"] for record in first_batch],
                [fresh_first["candidate_id"], fresh_second["candidate_id"]],
            )
            full_order = sync.select_publication_batch(records, state_path, 5)
            self.assertEqual(
                [record["candidate_id"] for record in full_order],
                [
                    fresh_first["candidate_id"],
                    fresh_second["candidate_id"],
                    failed_older["candidate_id"],
                    failed_newer["candidate_id"],
                    failed_many["candidate_id"],
                ],
            )

    def test_retry_classification_never_requires_evidence_for_published_or_fresh_rows(self):
        failed = exact_record("b" * 28)
        fresh = exact_record("c" * 28)
        published = exact_record("d" * 28)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            state_path = root / "pipeline.sqlite3"
            evidence_path = root / "spotify.json"
            evidence_path.write_text(
                json.dumps(
                    {
                        "records": [
                            spotify_evidence_record(failed["candidate_id"]),
                        ]
                    }
                ),
                encoding="utf-8",
            )
            attached = json.loads(json.dumps([failed]))
            sync.process.attach_verified_spotify_evidence(attached, evidence_path)
            connection = sync.process.open_pipeline_state(state_path)
            now = "2026-01-01T00:00:00+00:00"
            connection.execute(
                """INSERT INTO pipeline_items (
                candidate_id, manifest_fingerprint, batch_key, status,
                attempts, created_at, updated_at
                ) VALUES (?, ?, 'old-batch', 'failed', 2, ?, ?)""",
                (
                    failed["candidate_id"],
                    sync.process.canonical_fingerprint(attached[0]),
                    now,
                    now,
                ),
            )
            connection.execute(
                """INSERT INTO pipeline_items (
                candidate_id, manifest_fingerprint, batch_key, status,
                attempts, created_at, updated_at
                ) VALUES (?, ?, 'published-batch', 'published', 1, ?, ?)""",
                (published["candidate_id"], "f" * 64, now, now),
            )
            connection.execute(
                """CREATE TABLE published_manifest_bases (
                candidate_id TEXT PRIMARY KEY,
                base_fingerprint TEXT NOT NULL,
                source_manifest_fingerprint TEXT NOT NULL,
                verified_at TEXT NOT NULL
                )"""
            )
            connection.execute(
                "INSERT INTO published_manifest_bases VALUES (?, ?, ?, ?)",
                (
                    published["candidate_id"],
                    sync.process.canonical_fingerprint(published),
                    "f" * 64,
                    now,
                ),
            )
            connection.commit()
            connection.close()

            selected = sync.select_publication_batch(
                [failed, published, fresh],
                state_path,
                3,
                spotify_enrichment=evidence_path,
            )
            self.assertEqual(
                [record["candidate_id"] for record in selected],
                [fresh["candidate_id"], failed["candidate_id"]],
            )

    def test_bootstrap_merges_only_published_rows_with_current_fingerprint_and_acks(self):
        matching = exact_record("b" * 28)
        changed = exact_record("c" * 28)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "batch-old" / "pipeline-state.sqlite3"
            target_path = root / "unified" / "pipeline-state.sqlite3"
            source = sync.process.open_pipeline_state(source_path)
            now = dt.datetime.now(dt.timezone.utc).isoformat()
            for record in (matching, changed):
                source.execute(
                    """INSERT INTO pipeline_items (
                    candidate_id, manifest_fingerprint, batch_key, status,
                    rights_cleared_ack, human_made_cleared_ack, created_at, updated_at
                    ) VALUES (?, ?, ?, 'published', 1, 1, ?, ?)""",
                    (
                        record["candidate_id"],
                        sync.process.canonical_fingerprint(record),
                        sync.process.DEFAULT_BATCH_KEY,
                        now,
                        now,
                    ),
                )
            source.commit()
            source.close()
            changed["track"]["genre"] = "Changed after publication"
            counts = sync.bootstrap_published_pipeline_state(
                [matching, changed], target_path, search_root=root
            )
            self.assertEqual(counts["merged"], 1)
            self.assertEqual(counts["refused"], 1)
            target = sqlite3.connect(target_path)
            rows = target.execute("SELECT candidate_id FROM pipeline_items").fetchall()
            target.close()
            self.assertEqual(rows, [(matching["candidate_id"],)])

    def test_ingestion_bootstrap_reuses_only_full_inspection_with_same_fingerprint(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target_path = root / "unified" / "ingestion-state.sqlite3"
            source_path = root / "batch-old" / "ingestion-state.sqlite3"
            target = sync.ingest.open_state(target_path)
            source = sync.ingest.open_state(source_path)
            payload = json.dumps({"track": {}, "audio": {}})
            target.execute(
                """INSERT INTO candidates (
                candidate_id, fingerprint, payload_json, status, reasons_json, updated_at
                ) VALUES ('matching', 'same', ?, 'review', '[]', 'now')""",
                (payload,),
            )
            target.execute(
                """INSERT INTO candidates (
                candidate_id, fingerprint, payload_json, status, reasons_json, updated_at
                ) VALUES ('changed', 'new', ?, 'review', '[]', 'now')""",
                (payload,),
            )
            wav = json.dumps(
                {
                    "container": "RIFF",
                    "codec": "PCM",
                    "channels": 2,
                    "sample_rate": 48000,
                    "bits_per_sample": 16,
                    "byte_rate": 192000,
                    "data_size": 23040000,
                    "duration_seconds": 120,
                }
            )
            for candidate_id in ("matching", "changed"):
                source.execute(
                    """INSERT INTO candidates (
                    candidate_id, fingerprint, payload_json, status, reasons_json,
                    inspection_status, inspection_mode, wav_json, sha256, updated_at
                    ) VALUES (?, 'same', ?, 'exact', '[]', 'complete', 'full', ?, ?, 'now')""",
                    (candidate_id, payload, wav, "a" * 64),
                )
            target.commit()
            source.commit()
            target.close()
            source.close()
            counts = sync.bootstrap_full_ingestion_state(target_path, search_root=root)
            self.assertEqual(counts["merged"], 1)
            self.assertEqual(counts["refused"], 1)
            target = sqlite3.connect(target_path)
            rows = dict(target.execute("SELECT candidate_id, inspection_status FROM candidates"))
            target.close()
            self.assertEqual(rows, {"matching": "complete", "changed": "pending"})


class EvidenceTests(unittest.TestCase):
    def test_spotify_history_bootstrap_keeps_only_current_unambiguous_structured_rows(self):
        accepted_key = "b" * 28
        review_key = "c" * 28
        ambiguous_key = "d" * 28
        stale_key = "e" * 28
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first = root / "batch-one" / "spotify-enrichment" / "enriched-tracks.json"
            second = root / "batch-two" / "spotify-enrichment" / "enriched-tracks.json"
            malformed = root / "batch-bad" / "spotify-enrichment" / "enriched-tracks.json"
            cumulative = root / "unified" / "spotify-evidence.json"
            first.parent.mkdir(parents=True)
            second.parent.mkdir(parents=True)
            malformed.parent.mkdir(parents=True)
            accepted = spotify_evidence_record(accepted_key)
            duplicate = json.loads(json.dumps(accepted))
            duplicate["inputIndex"] = 99
            duplicate["cache"] = "hit"
            first.write_text(
                json.dumps(
                    {
                        "records": [
                            accepted,
                            spotify_evidence_record(review_key, disposition="review"),
                            spotify_evidence_record(ambiguous_key, title="First title"),
                            spotify_evidence_record(stale_key),
                            spotify_evidence_record("not-a-candidate"),
                        ]
                    }
                ),
                encoding="utf-8",
            )
            second.write_text(
                json.dumps(
                    {
                        "records": [
                            duplicate,
                            spotify_evidence_record(ambiguous_key, title="Second title"),
                        ]
                    }
                ),
                encoding="utf-8",
            )
            malformed.write_text("not json", encoding="utf-8")

            counts = sync.bootstrap_cumulative_spotify(
                cumulative,
                search_root=root,
                allowed_record_keys={accepted_key, review_key, ambiguous_key},
            )
            payload = json.loads(cumulative.read_text(encoding="utf-8"))

        self.assertEqual(counts["records"], 2)
        self.assertEqual(counts["accepted"], 1)
        self.assertEqual(counts["review"], 1)
        self.assertEqual(counts["duplicates"], 1)
        self.assertEqual(counts["ambiguous"], 1)
        self.assertEqual(counts["staleRecords"], 1)
        self.assertEqual(counts["invalidRecords"], 1)
        self.assertEqual(counts["invalidSources"], 1)
        self.assertEqual(
            {record["recordKey"] for record in payload["records"]},
            {accepted_key, review_key},
        )
        self.assertNotIn("cache", payload["records"][0])
        self.assertNotIn("inputIndex", payload["records"][0])
        self.assertNotIn(accepted_key, json.dumps(counts))

    def test_current_spotify_batch_overrides_history_and_invalidates_newer_failures(self):
        replaced_key = "b" * 28
        unavailable_key = "c" * 28
        ambiguous_key = "d" * 28
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            cumulative = root / "spotify-evidence.json"
            current = root / "spotify-enrichment" / "enriched-tracks.json"
            current.parent.mkdir(parents=True)
            cumulative.write_text(
                json.dumps(
                    {
                        "records": [
                            spotify_evidence_record(replaced_key, title="Old title"),
                            spotify_evidence_record(unavailable_key),
                        ]
                    }
                ),
                encoding="utf-8",
            )
            unavailable = spotify_evidence_record(unavailable_key)
            unavailable["disposition"] = "unavailable"
            unavailable["spotify"] = None
            first_conflict = spotify_evidence_record(ambiguous_key, title="First")
            second_conflict = spotify_evidence_record(ambiguous_key, title="Second")
            current.write_text(
                json.dumps(
                    {
                        "records": [
                            spotify_evidence_record(replaced_key, title="Current title"),
                            unavailable,
                            first_conflict,
                            second_conflict,
                        ]
                    }
                ),
                encoding="utf-8",
            )

            counts = sync.merge_cumulative_spotify(current, cumulative)
            payload = json.loads(cumulative.read_text(encoding="utf-8"))

        self.assertEqual(counts["records"], 1)
        self.assertEqual(counts["accepted"], 1)
        self.assertEqual(counts["currentInvalid"], 1)
        self.assertEqual(counts["currentAmbiguous"], 1)
        self.assertEqual(payload["records"][0]["spotify"]["title"], "Current title")

    def test_three_historical_acceptances_survive_process_revalidation(self):
        records = [exact_record(character * 28) for character in ("b", "c", "d")]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "old" / "spotify-enrichment" / "enriched-tracks.json"
            cumulative = root / "unified" / "spotify-evidence.json"
            source.parent.mkdir(parents=True)
            source.write_text(
                json.dumps(
                    {
                        "records": [
                            spotify_evidence_record(record["candidate_id"])
                            for record in records
                        ]
                    }
                ),
                encoding="utf-8",
            )
            counts = sync.bootstrap_cumulative_spotify(
                cumulative,
                search_root=root,
                allowed_record_keys=(record["candidate_id"] for record in records),
            )
            copied = json.loads(json.dumps(records))
            attached = sync.process.attach_verified_spotify_evidence(copied, cumulative)

        self.assertEqual(counts["accepted"], 3)
        self.assertEqual(attached["attached"], 3)

    def test_evidence_is_explicit_and_bound_to_the_selection_hash(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "rights.json"
            selection_hash = hashlib.sha256(b"selection").hexdigest()
            path.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "kind": "rights_clearance",
                        "approved": True,
                        "selectionSha256": selection_hash,
                        "selectionCount": 3,
                        "reviewer": "Catalogue team",
                        "reviewedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
                    }
                ),
                encoding="utf-8",
            )
            sync.validate_evidence(path, "rights_clearance", selection_hash, 3)
            with self.assertRaisesRegex(sync.SyncError, "rights_clearance_evidence_invalid"):
                sync.validate_evidence(path, "rights_clearance", "0" * 64, 3)

    def test_owner_attestation_derives_private_evidence_for_each_exact_selection(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            drive_root = "ABCDEFGHIJKLMNO"
            workbook_id = "S" * 20
            seed = root / "seed.json"
            owner = root / "catalog-owner-attestation.json"
            rights = root / "rights.json"
            human = root / "human.json"
            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": drive_root,
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            owner.write_text(
                json.dumps(owner_attestation_payload(drive_root, workbook_id)),
                encoding="utf-8",
            )
            config = {
                "catalogue_source_url": "https://lofi-records.netlify.app/#/catalog",
                "workbook_source_url": f"https://docs.google.com/spreadsheets/d/{workbook_id}/edit?usp=sharing",
                "drive_seed": seed,
                "catalog_owner_attestation": owner,
                "rights_evidence": rights,
                "human_evidence": human,
            }

            first_hash = hashlib.sha256(b"first-selection").hexdigest()
            result = sync.authorize_publication(config, first_hash, 3)
            self.assertEqual(result["source"], "catalog_owner_attestation")
            self.assertTrue(result["selectionEvidenceDerived"])
            sync.validate_evidence(rights, "rights_clearance", first_hash, 3)
            sync.validate_evidence(human, "human_made_editorial_review", first_hash, 3)

            derived = rights.read_text(encoding="utf-8") + human.read_text(encoding="utf-8")
            self.assertNotIn(drive_root, derived)
            self.assertNotIn(workbook_id, derived)
            self.assertNotIn("lofi-records.netlify.app", derived)

            second_hash = hashlib.sha256(b"second-selection").hexdigest()
            second = sync.authorize_publication(config, second_hash, 2)
            self.assertTrue(second["selectionEvidenceDerived"])
            sync.validate_evidence(rights, "rights_clearance", second_hash, 2)
            sync.validate_evidence(human, "human_made_editorial_review", second_hash, 2)
            self.assertNotEqual(
                json.loads(rights.read_text(encoding="utf-8"))["selectionSha256"],
                first_hash,
            )

    def test_owner_attestation_is_invalidated_by_any_source_scope_change(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            seed = root / "seed.json"
            owner = root / "catalog-owner-attestation.json"
            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": "PQRSTUVWXYZabcd",
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            owner.write_text(json.dumps(owner_attestation_payload()), encoding="utf-8")
            config = {
                "catalogue_source_url": "https://lofi-records.netlify.app/#/catalog",
                "workbook_source_url": "https://docs.google.com/spreadsheets/d/" + "S" * 20,
                "drive_seed": seed,
            }
            with self.assertRaisesRegex(sync.SyncError, "catalog_owner_attestation_scope_invalid"):
                sync.validate_catalog_owner_attestation(owner, config)

    def test_derived_evidence_cannot_outlive_its_attested_scope(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            seed = root / "seed.json"
            owner = root / "catalog-owner-attestation.json"
            rights = root / "rights.json"
            human = root / "human.json"
            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": "ABCDEFGHIJKLMNO",
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            owner.write_text(json.dumps(owner_attestation_payload()), encoding="utf-8")
            config = {
                "catalogue_source_url": "https://lofi-records.netlify.app/#/catalog",
                "workbook_source_url": "https://docs.google.com/spreadsheets/d/" + "S" * 20,
                "drive_seed": seed,
                "catalog_owner_attestation": owner,
                "rights_evidence": rights,
                "human_evidence": human,
            }
            selection_hash = hashlib.sha256(b"selection").hexdigest()
            sync.authorize_publication(config, selection_hash, 1)

            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": "PQRSTUVWXYZabcd",
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(sync.SyncError, "catalog_owner_attestation_scope_invalid"):
                sync.authorize_publication(config, selection_hash, 1)

            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": "ABCDEFGHIJKLMNO",
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            config["workbook_source_url"] = "https://docs.google.com/spreadsheets/d/" + "T" * 20
            with self.assertRaisesRegex(sync.SyncError, "catalog_owner_attestation_scope_invalid"):
                sync.validate_catalog_owner_attestation(owner, config)

    def test_owner_attestation_requires_all_rights_and_non_ai_claims(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            seed = root / "seed.json"
            owner = root / "catalog-owner-attestation.json"
            seed.write_text(
                json.dumps(
                    {
                        "schemaVersion": 1,
                        "sourceFolderId": "ABCDEFGHIJKLMNO",
                        "releaseFolders": [],
                    }
                ),
                encoding="utf-8",
            )
            payload = owner_attestation_payload()
            payload["claims"]["humanMadeNoGenerativeAI"] = False
            owner.write_text(json.dumps(payload), encoding="utf-8")
            config = {
                "catalogue_source_url": "https://lofi-records.netlify.app/#/catalog",
                "workbook_source_url": "https://docs.google.com/spreadsheets/d/" + "S" * 20,
                "drive_seed": seed,
            }
            with self.assertRaisesRegex(sync.SyncError, "catalog_owner_attestation_invalid"):
                sync.validate_catalog_owner_attestation(owner, config)

    def test_publish_command_never_gets_attestations_from_configuration(self):
        config = {
            "workbook": Path("private.xlsx"),
            "orchard": Path("orchard.xlsx"),
            "work_directory": Path("private"),
            "inspection_batch": 5,
            "release_batch": 2,
            "maximum_run_minutes": 45,
        }
        command = sync.build_ingest_command(
            config, Path("inventory.json"), apply=True, inspect_full=True
        )
        self.assertIn("--inspect", command)
        self.assertIn("full", command)
        self.assertIn("--release-batch-size", command)
        self.assertIn("--max-inspection-seconds", command)
        self.assertNotIn("--rights-cleared", command)
        self.assertNotIn("--human-made-cleared", command)


if __name__ == "__main__":
    unittest.main()
