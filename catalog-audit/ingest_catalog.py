#!/usr/bin/env python3
"""Build and verify a private, resumable Symbiome catalogue manifest.

This command deliberately does not publish masters and does not copy audio into
the repository.  It joins the private catalogue workbook, Drive audio/artwork
references and The Orchard's Spotify mapping, then stores its per-track state in
``catalog-audit/private`` (which is ignored by Git).

Network access is opt-in.  ``range`` inspection reads only the beginning of a
WAV; ``full`` streams one WAV at a time to a temporary file to calculate its
SHA-256 and removes that file immediately afterwards.
"""

from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import io
import json
import math
import os
import re
import sqlite3
import struct
import sys
import tempfile
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import warnings
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence


PRIVATE_DIRECTORY = Path(__file__).resolve().parent / "private"
DRIVE_FOLDER_MIME = "application/vnd.google-apps.folder"
WAV_MIME_TYPES = {"audio/wav", "audio/x-wav", "audio/wave", "audio/vnd.wave"}
MP3_MIME_TYPES = {"audio/mpeg", "audio/mp3", "audio/x-mp3"}
AUDIO_FOLDER_NAMES = {"wav", "wave", "music", "musique", ".wav"}
VERSION_TOKENS = {
    "acoustic",
    "clean",
    "demo",
    "edit",
    "explicit",
    "instrumental",
    "karaoke",
    "live",
    "radio",
    "remaster",
    "remastered",
    "remix",
    "slowed",
    "sped",
    "tribute",
}
PARASITE_TOKENS = {"preview", "snippet", "sample", "test", "bounce test", "copy of"}
DRIVE_ID_PATTERN = re.compile(r"[A-Za-z0-9_-]{15,}")
SPOTIFY_ID_PATTERN = re.compile(r"(?:spotify:track:|open\.spotify\.com/track/)([A-Za-z0-9]{22})")
DIRECT_PUBLICATION_BLOCKERS = frozenset(
    {
        "audio_match_missing",
        "audio_match_ambiguous",
        "audio_file_claimed_by_multiple_tracks",
        "unmatched_audio_file",
        "source_snapshot_missing",
    }
)
STRICT_CENTRAL_MP3_MATCH_KINDS = frozenset(
    {
        "central_unique_isrc",
        "central_unique_upc_title",
        "central_unique_artist_title",
        "central_unique_release_title",
    }
)
CENTRAL_AUDIO_PIN_RULES = (
    "central_unique_isrc",
    "central_unique_upc_title",
    "central_unique_artist_title",
    "central_unique_release_title",
    "central_globally_unique_exact_title",
)


@dataclasses.dataclass(frozen=True)
class Track:
    source_row: int
    release: str
    upc: str
    title: str
    artists: tuple[str, ...]
    duration_seconds: float | None
    isrc: str
    genre: str
    subgenre: str

    @property
    def normalized_release(self) -> str:
        return normalize_text(self.release)

    @property
    def normalized_title(self) -> str:
        return normalize_text(self.title)

    @property
    def stable_id(self) -> str:
        raw = self.isrc or "|".join(
            [self.upc, self.normalized_release, self.normalized_title, *(normalize_text(value) for value in self.artists)]
        )
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


@dataclasses.dataclass(frozen=True)
class DriveAudio:
    file_id: str
    name: str
    folder_id: str = ""
    folder_path: str = ""
    mime_type: str = ""
    size_bytes: int | None = None
    modified_time: str = ""

    @property
    def stable_id(self) -> str:
        return hashlib.sha256(f"drive:{self.file_id}".encode("utf-8")).hexdigest()[:24]


@dataclasses.dataclass(frozen=True)
class ReleaseAudio:
    upc: str
    release: str
    folder_id: str
    files: tuple[DriveAudio, ...]


@dataclasses.dataclass(frozen=True)
class CentralAudioPin:
    """One central file bound to one exact workbook row.

    ``source_row`` is intentionally part of the binding. A release-wide audio
    group is not precise enough here: adding such a group can make an
    already-good neighbour ambiguous when a larger central snapshot arrives.
    """

    source_row: int
    track_stable_id: str
    audio: DriveAudio
    match_kind: str


@dataclasses.dataclass(frozen=True)
class Cover:
    upc: str
    file_id: str
    quality: str
    is_square: bool


@dataclasses.dataclass(frozen=True)
class OrchardTrack:
    upc: str
    release: str
    title: str
    artist: str
    spotify_id: str
    spotify_duration_seconds: float | None
    active: bool


@dataclasses.dataclass
class Candidate:
    candidate_id: str
    track: Track | None
    audio: DriveAudio | None
    audio_match_score: float | None
    audio_match_kind: str
    spotify_id: str
    spotify_duration_seconds: float | None
    spotify_match_kind: str
    cover: Cover | None
    status: str
    reasons: list[str]
    fingerprint: str


@dataclasses.dataclass(frozen=True)
class WavInfo:
    container: str
    codec: str
    channels: int
    sample_rate: int
    bits_per_sample: int
    byte_rate: int
    data_size: int
    duration_seconds: float


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value).casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_upc(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    digits = re.sub(r"\D", "", clean(value))
    return digits.lstrip("0") or digits


def normalize_isrc(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value)).upper()


def parse_expected_duration(value: Any) -> float | None:
    if value is None or clean(value) == "":
        return None
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    if isinstance(value, (dt.datetime, dt.time)):
        return float(value.hour * 3600 + value.minute * 60 + value.second) + value.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        numeric = float(value)
        # Excel stores clock values as a fraction of a day.
        return numeric * 86400 if 0 <= numeric < 1 else numeric
    match = re.fullmatch(r"(?:(\d+):)?(\d+):([0-5]\d)(?:[.,](\d+))?", clean(value))
    if match:
        hours = int(match.group(1) or 0)
        minutes = int(match.group(2))
        seconds = int(match.group(3))
        fraction = float(f"0.{match.group(4)}") if match.group(4) else 0
        return hours * 3600 + minutes * 60 + seconds + fraction
    return None


def extract_drive_id(value: Any, *, folder: bool | None = None) -> str:
    text = clean(value)
    if not text:
        return ""
    patterns: list[str] = []
    if folder is not False:
        patterns.append(r"/folders/([^?/#]+)")
    if folder is not True:
        patterns.extend([r"/file/d/([^/]+)", r"/d/([^/]+)", r"[?&]id=([^&#]+)"])
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        candidate = re.sub(r"\s+", "", urllib.parse.unquote(match.group(1)))
        id_match = DRIVE_ID_PATTERN.search(candidate)
        if id_match:
            return id_match.group(0)
    if DRIVE_ID_PATTERN.fullmatch(re.sub(r"\s+", "", text)):
        return re.sub(r"\s+", "", text)
    return ""


def extract_spotify_id(*values: Any) -> str:
    for value in values:
        candidate = clean(value)
        if re.fullmatch(r"[A-Za-z0-9]{22}", candidate):
            return candidate
        match = SPOTIFY_ID_PATTERN.search(candidate)
        if match:
            return match.group(1)
    return ""


def parse_audio_entries(value: Any, *, fallback_name: str = "") -> tuple[DriveAudio, ...]:
    text = clean(value)
    if not text or text.casefold() in {"#n/a", "n/a", "not found"}:
        return ()
    entries: list[DriveAudio] = []
    seen: set[str] = set()
    lines = [line.strip() for line in re.split(r"[\r\n]+", text) if line.strip()]
    for line in lines:
        file_id = extract_drive_id(line, folder=False)
        if not file_id or file_id in seen:
            continue
        prefix = line.split("http", 1)[0].rstrip(" :-\t")
        wav_match = re.search(r"([^:\r\n]+?\.wav)\s*$", prefix, flags=re.IGNORECASE)
        name = clean(wav_match.group(1) if wav_match else fallback_name)
        if not name:
            name = f"unknown-{hashlib.sha256(file_id.encode()).hexdigest()[:8]}.wav"
        if not name.casefold().endswith(".wav"):
            name = f"{name}.wav"
        entries.append(DriveAudio(file_id=file_id, name=name))
        seen.add(file_id)
    return tuple(entries)


def header_map(sheet: Any) -> dict[str, int]:
    return {clean(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1) if clean(sheet.cell(1, column).value)}


def cell_value(sheet: Any, row: int, headers: Mapping[str, int], *names: str) -> Any:
    for name in names:
        column = headers.get(name)
        if column:
            return sheet.cell(row, column).value
    return None


def cell_link_or_value(cell: Any) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        return clean(cell.hyperlink.target)
    return clean(cell.value)


def cell_display_text(cell: Any) -> str:
    """Return the human label from an exported hyperlink formula.

    Google Sheets XLSX exports sometimes keep ``HYPERLINK(url, label)`` as a
    formula without a cached display value.  Catalogue joins must compare the
    release label, never the formula or Drive URL.
    """

    value = clean(cell.value)
    match = re.fullmatch(
        r'=HYPERLINK\(\s*"(?:[^"]|"")*"\s*[,;]\s*"((?:[^"]|"")*)"\s*\)',
        value,
        flags=re.IGNORECASE,
    )
    return match.group(1).replace('""', '"') if match else value


def load_workbook_sources(workbook_path: Path) -> tuple[list[Track], list[ReleaseAudio], dict[str, Cover]]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as error:
        raise RuntimeError("openpyxl is required to read All DATA.xlsx (python -m pip install openpyxl).") from error

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)

    required = {"Publishing catalogue", "Cover Album"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing sheet(s): {', '.join(sorted(missing))}")

    publishing = workbook["Publishing catalogue"]
    headers = header_map(publishing)
    tracks: list[Track] = []
    for row in range(2, publishing.max_row + 1):
        title = clean(cell_value(publishing, row, headers, "Track Title"))
        release_column = headers.get("Album Name (Drive Link)")
        release = cell_display_text(publishing.cell(row, release_column)) if release_column else ""
        if not title and not release:
            continue
        artists = tuple(
            clean(cell_value(publishing, row, headers, f"Artist {number}"))
            for number in range(1, 5)
            if clean(cell_value(publishing, row, headers, f"Artist {number}"))
        )
        tracks.append(
            Track(
                source_row=row,
                release=release,
                upc=normalize_upc(cell_value(publishing, row, headers, "UPC")),
                title=title,
                artists=artists,
                duration_seconds=parse_expected_duration(cell_value(publishing, row, headers, "Track Time")),
                isrc=normalize_isrc(cell_value(publishing, row, headers, "ISRC")),
                genre=clean(cell_value(publishing, row, headers, "Genre")),
                subgenre=clean(cell_value(publishing, row, headers, "Sous Genre", "Subgenre")),
            )
        )

    covers_sheet = workbook["Cover Album"]
    cover_headers = header_map(covers_sheet)
    covers: dict[str, Cover] = {}
    for row in range(2, covers_sheet.max_row + 1):
        upc = normalize_upc(cell_value(covers_sheet, row, cover_headers, "UPC"))
        link_column = cover_headers.get("Link Artwork")
        if not upc or not link_column:
            continue
        file_id = extract_drive_id(cell_link_or_value(covers_sheet.cell(row, link_column)), folder=False)
        if not file_id:
            continue
        quality = clean(cell_value(covers_sheet, row, cover_headers, "Nom Fichier"))
        square = "carr" in normalize_text(quality) or bool(re.search(r"\b(\d{3,5})\s*x\s*\1\b", quality))
        covers.setdefault(upc, Cover(upc=upc, file_id=file_id, quality=quality, is_square=square))

    release_groups: dict[tuple[str, str], dict[str, Any]] = {}
    for sheet_name in ("Release Drive link + Audio", "Release Drive + Audio 2"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        # The sheets contain two independent blocks. A:H is per publishing row;
        # I:M is the curated release/folder/audio block. Parse both and de-dupe.
        for row in range(2, sheet.max_row + 1):
            for release_column, upc_column, folder_column, audio_column, track_column in (
                (2, 4, None, 6, 5),
                (9, 10, 11, 13, None),
            ):
                release = clean(sheet.cell(row, release_column).value)
                upc = normalize_upc(sheet.cell(row, upc_column).value)
                if not release and not upc:
                    continue
                folder_id = ""
                if folder_column:
                    folder_id = extract_drive_id(cell_link_or_value(sheet.cell(row, folder_column)), folder=True)
                fallback = clean(sheet.cell(row, track_column).value) if track_column else ""
                files = parse_audio_entries(sheet.cell(row, audio_column).value, fallback_name=fallback)
                key = (upc, normalize_text(release))
                group = release_groups.setdefault(key, {"upc": upc, "release": release, "folder_id": folder_id, "files": {}})
                if folder_id and not group["folder_id"]:
                    group["folder_id"] = folder_id
                for audio in files:
                    group["files"].setdefault(audio.file_id, audio)

    release_audio = [
        ReleaseAudio(
            upc=group["upc"],
            release=group["release"],
            folder_id=group["folder_id"],
            files=tuple(group["files"].values()),
        )
        for group in release_groups.values()
    ]
    workbook.close()
    return tracks, release_audio, covers


def load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    extension = path.suffix.casefold()
    if extension == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(payload, dict):
            payload = payload.get("values", payload.get("rows", payload.get("files", [])))
        if not isinstance(payload, list):
            raise ValueError(f"{path}: JSON must contain a list or a values/rows/files list.")
        if payload and isinstance(payload[0], list):
            headers = [clean(value) for value in payload[0]]
            return [dict(zip(headers, row)) for row in payload[1:] if any(clean(value) for value in row)]
        return [dict(row) for row in payload if isinstance(row, dict)]
    if extension in {".csv", ".tsv"}:
        text = path.read_text(encoding="utf-8-sig")
        delimiter = "\t" if extension == ".tsv" else (";" if text.partition("\n")[0].count(";") > text.partition("\n")[0].count(",") else ",")
        return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))
    if extension == ".xlsx":
        try:
            import openpyxl  # type: ignore[import-not-found]
        except ImportError as error:
            raise RuntimeError("openpyxl is required for XLSX inputs.") from error
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        result = [dict(zip(headers, row)) for row in rows if any(clean(value) for value in row)]
        workbook.close()
        return result
    raise ValueError(f"Unsupported private export: {path.suffix}. Use XLSX, CSV, TSV or JSON.")


def row_value(row: Mapping[str, Any], *names: str) -> Any:
    folded = {normalize_text(key): value for key, value in row.items()}
    for name in names:
        key = normalize_text(name)
        if key in folded:
            return folded[key]
    return None


def load_orchard(path: Path | None) -> list[OrchardTrack]:
    if path is None:
        return []
    result: list[OrchardTrack] = []
    for row in load_tabular_rows(path):
        spotify_id = extract_spotify_id(row_value(row, "Spotify URI"), row_value(row, "URL"))
        duration_ms_raw = row_value(row, "Spotify Duration ms", "Duration ms", "duration_ms")
        try:
            spotify_duration = float(duration_ms_raw) / 1000 if clean(duration_ms_raw) else None
        except (TypeError, ValueError):
            spotify_duration = None
        active_value = normalize_text(row_value(row, "Is Active (Yes / No)", "Is Active", "Active"))
        result.append(
            OrchardTrack(
                upc=normalize_upc(row_value(row, "UPC")),
                release=clean(row_value(row, "Release Name")),
                title=clean(row_value(row, "Track Name")),
                artist=clean(row_value(row, "Artist Name")),
                spotify_id=spotify_id,
                spotify_duration_seconds=spotify_duration,
                active=active_value not in {"no", "false", "0", "inactive"},
            )
        )
    return result


def load_identity_info(workbook_path: Path) -> list[OrchardTrack]:
    """Load the workbook's refreshed distributor identity fallback.

    `track Id` is accepted as Spotify evidence only when it is an actual
    22-character Spotify ID/URI/URL. Numeric distributor IDs remain empty and
    therefore keep publication closed until a verified Spotify mapping exists.
    """

    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as error:
        raise RuntimeError("openpyxl is required for Identity Info.") from error
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Identity Info" not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook["Identity Info"]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(rows)]
    except StopIteration:
        workbook.close()
        return []
    result: list[OrchardTrack] = []
    for values in rows:
        row = dict(zip(headers, values))
        upc = normalize_upc(row_value(row, "AlbumUPC"))
        release = clean(row_value(row, "AlbumTitle"))
        title = clean(row_value(row, "TrackTitle"))
        artist = clean(row_value(row, "TrackArtist"))
        if not upc or not release or not title or not artist:
            continue
        result.append(
            OrchardTrack(
                upc=upc,
                release=release,
                title=title,
                artist=artist,
                spotify_id=extract_spotify_id(row_value(row, "track Id", "Spotify URI", "Spotify URL")),
                spotify_duration_seconds=parse_expected_duration(row_value(row, "Duration")),
                active=normalize_text(row_value(row, "Check")) not in {"3", "inactive", "no", "false"},
            )
        )
    workbook.close()
    return result


def merge_orchard_rows(primary: Sequence[OrchardTrack], fallback: Sequence[OrchardTrack]) -> list[OrchardTrack]:
    """Prefer curated Orchard/Spotify rows and fill only absent strict keys."""

    result: list[OrchardTrack] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in (*primary, *fallback):
        key = (row.upc, normalize_text(row.release), normalize_text(row.title), normalize_text(row.artist))
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def load_drive_inventory(path: Path | None) -> tuple[dict[str, list[DriveAudio]], dict[str, list[dict[str, Any]]]]:
    """Return files grouped by release folder and raw children grouped by parent.

    Accepted connector-export fields are intentionally flexible: ``id``,
    ``name``, ``mimeType``, ``parents``, ``path`` and ``release_folder_id``.
    The private export itself is never copied outside the ignored directory.
    """

    if path is None:
        return {}, {}
    rows = load_tabular_rows(path)
    by_release: dict[str, list[DriveAudio]] = defaultdict(list)
    by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        marker = normalize_text(row_value(row, "inventoryMarker", "inventory_marker"))
        if marker in {"1", "true", "yes"}:
            parent = clean(row_value(row, "parent_id", "parentId"))
            if parent:
                # An explicit empty-folder marker prevents live discovery from
                # falling back to OAuth when a complete public snapshot says
                # the folder simply has no relevant children.
                by_parent.setdefault(parent, [])
            continue
        file_id = extract_drive_id(row_value(row, "id", "file_id", "webViewLink"), folder=False)
        if not file_id:
            continue
        name = clean(row_value(row, "name", "file_name"))
        mime_type = clean(row_value(row, "mimeType", "mime_type"))
        parents_raw = row_value(row, "parents", "parent_id")
        if isinstance(parents_raw, list):
            parents = [clean(value) for value in parents_raw]
        else:
            parents = [value for value in re.split(r"[,;\s]+", clean(parents_raw)) if value]
        normalized = dict(row)
        normalized.update({"id": file_id, "name": name, "mimeType": mime_type, "parents": parents})
        for parent in parents:
            by_parent[parent].append(normalized)
        release_folder_id = extract_drive_id(row_value(row, "release_folder_id", "releaseFolderId"), folder=True) or clean(
            row_value(row, "release_folder_id", "releaseFolderId")
        )
        is_wav = name.casefold().endswith(".wav") or mime_type.casefold() in WAV_MIME_TYPES
        if release_folder_id and is_wav:
            by_release[release_folder_id].append(drive_audio_from_row(normalized))
    return dict(by_release), dict(by_parent)


def load_central_drive_inventory(
    path: Path | None,
) -> tuple[list[DriveAudio], list[dict[str, Any]], bool]:
    """Load an additive flat catalogue snapshot from the central Drive folders.

    Connector snapshots may be partial (notably when a provider page is
    capped).  A partial snapshot is still useful for positive matches, but it
    can never imply that an older mapping was removed.
    """

    if path is None or not path.is_file():
        return [], [], False
    rows = load_tabular_rows(path)
    audio: dict[str, DriveAudio] = {}
    artwork: dict[str, dict[str, Any]] = {}
    for row in rows:
        file_id = extract_drive_id(row_value(row, "id", "file_id", "webViewLink"), folder=False)
        if not file_id:
            continue
        name = clean(row_value(row, "name", "file_name"))
        mime_type = clean(row_value(row, "mimeType", "mime_type"))
        central_kind = normalize_text(row_value(row, "central_kind", "centralKind", "kind"))
        descriptor = audio_source_descriptor({"name": name, "mime_type": mime_type})
        is_audio = descriptor is not None
        is_artwork = mime_type.casefold().startswith("image/") or Path(name).suffix.casefold() in {
            ".jpg",
            ".jpeg",
            ".png",
            ".webp",
        }
        if is_audio and central_kind in {"", "audio"}:
            audio.setdefault(
                file_id,
                drive_audio_from_row(
                    row,
                    folder_id=clean(row_value(row, "parent_id", "parentId")),
                    folder_path=clean(row_value(row, "path", "folder_path", "folderPath")),
                ),
            )
        elif is_artwork and central_kind in {"", "artwork", "cover"}:
            artwork.setdefault(file_id, dict(row))
    complete = False
    if path.suffix.casefold() == ".json":
        try:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
            complete = isinstance(payload, dict) and payload.get("complete") is True
        except (OSError, json.JSONDecodeError):
            complete = False
    return list(audio.values()), list(artwork.values()), complete


def _audio_stem(value: str) -> str:
    stem = re.sub(r"\.(?:wav|wave|mp3)$", "", clean(value), flags=re.IGNORECASE)
    stem = re.sub(r"^\s*\d{1,3}\s*[._-]+\s*", "", stem)
    return normalize_text(stem)


def _compact_identifier(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())


def audio_source_descriptor(value: DriveAudio | Mapping[str, Any]) -> tuple[str, str] | None:
    """Return the canonical format/MIME pair for supported owner sources.

    Extension and declared MIME are independent evidence.  If both are known
    and disagree, the source fails closed instead of letting a renamed file
    enter the direct lane.
    """

    if isinstance(value, DriveAudio):
        name = value.name
        raw_mimes = [value.mime_type]
    else:
        name = clean(value.get("name"))
        raw_mimes = [
            clean(value.get("source_mime_type")),
            clean(value.get("mime_type")),
        ]
    extension = Path(name).suffix.casefold()
    extension_format = (
        "wav" if extension in {".wav", ".wave"} else "mp3" if extension == ".mp3" else None
    )
    mime_formats: set[str] = set()
    for raw_mime in raw_mimes:
        normalized_mime = raw_mime.split(";", 1)[0].strip().casefold()
        if normalized_mime in WAV_MIME_TYPES:
            mime_formats.add("wav")
        elif normalized_mime in MP3_MIME_TYPES:
            mime_formats.add("mp3")
    if len(mime_formats) > 1:
        return None
    mime_format = next(iter(mime_formats), None)
    if extension_format and mime_format and extension_format != mime_format:
        return None
    source_format = extension_format or mime_format
    if source_format is None:
        return None
    return (source_format, "audio/wav" if source_format == "wav" else "audio/mpeg")


def _unique_audio_from_files(track: Track, files: Sequence[DriveAudio]) -> DriveAudio | None:
    scored = sorted(
        (
            (audio, *filename_match_score(audio, track))
            for audio in files
            if not incompatible_version(audio.name, track.title) and not is_suspicious_audio_name(audio.name)
        ),
        key=lambda item: (-item[1], normalize_text(item[0].name), item[0].file_id),
    )
    if not scored or scored[0][1] < 90:
        return None
    if len(scored) > 1 and scored[1][1] >= scored[0][1] - 2:
        return None
    return scored[0][0]


def _existing_unique_audio(track: Track, release_audio: Sequence[ReleaseAudio]) -> DriveAudio | None:
    return _unique_audio_from_files(track, match_release_audio(track, release_audio))


def merge_central_audio_mappings(
    tracks: Sequence[Track],
    release_audio: Sequence[ReleaseAudio],
    central_audio: Sequence[DriveAudio],
) -> tuple[list[ReleaseAudio], dict[str, int]]:
    """Fill missing audio only from one-to-one deterministic central matches.

    Accepted signals are a unique ISRC embedded in the filename, a unique
    UPC+title match, or an exact filename key that identifies one workbook row.
    A file matching several rows, several files matching one row, suspicious
    versions and already mapped rows are all left untouched.
    """

    by_id = {track.stable_id: track for track in tracks}
    title_keys: dict[str, set[str]] = defaultdict(set)
    isrc_index: dict[str, set[str]] = defaultdict(set)
    upc_index: dict[str, list[Track]] = defaultdict(list)
    for track in tracks:
        title = track.normalized_title
        artists = {normalize_text(artist) for artist in track.artists if normalize_text(artist)}
        if track.artists:
            artists.add(normalize_text(" ".join(track.artists)))
        keys = {title}
        for artist in artists:
            keys.update({f"{artist} {title}".strip(), f"{title} {artist}".strip()})
        if track.normalized_release:
            keys.update(
                {
                    f"{track.normalized_release} {title}".strip(),
                    f"{title} {track.normalized_release}".strip(),
                }
            )
        if track.upc:
            keys.update({f"{track.upc} {title}".strip(), f"{title} {track.upc}".strip()})
            upc_index[track.upc].append(track)
        compact_isrc = _compact_identifier(track.isrc)
        if compact_isrc:
            isrc_index[compact_isrc].add(track.stable_id)
            normalized_isrc = normalize_text(track.isrc)
            keys.update({f"{normalized_isrc} {title}".strip(), f"{title} {normalized_isrc}".strip()})
        for key in keys:
            if key:
                title_keys[key].add(track.stable_id)

    existing_audio_ids = {audio.file_id for group in release_audio for audio in group.files}
    groups_by_both: dict[tuple[str, str], list[DriveAudio]] = defaultdict(list)
    groups_by_upc: dict[str, list[DriveAudio]] = defaultdict(list)
    groups_by_release: dict[str, list[DriveAudio]] = defaultdict(list)
    for group in release_audio:
        normalized_release = normalize_text(group.release)
        groups_by_both[(group.upc, normalized_release)].extend(group.files)
        if group.upc:
            groups_by_upc[group.upc].extend(group.files)
        if normalized_release:
            groups_by_release[normalized_release].extend(group.files)
    already_mapped: set[str] = set()
    for track in tracks:
        files = groups_by_both.get((track.upc, track.normalized_release))
        if not files and track.upc:
            files = groups_by_upc.get(track.upc)
        if not files:
            files = groups_by_release.get(track.normalized_release, [])
        unique_files = list({audio.file_id: audio for audio in files}.values())
        if _unique_audio_from_files(track, unique_files) is not None:
            already_mapped.add(track.stable_id)
    claims_by_track: dict[str, list[DriveAudio]] = defaultdict(list)
    ambiguous_files = 0
    rejected_files = 0
    for audio in central_audio:
        if audio.file_id in existing_audio_ids:
            continue
        if is_suspicious_audio_name(audio.name):
            rejected_files += 1
            continue
        stem = _audio_stem(audio.name)
        compact_stem = _compact_identifier(Path(audio.name).stem)
        candidates = set(title_keys.get(stem, set()))

        # ISRC is globally unique when the source itself has no duplicate.
        for isrc, track_ids in isrc_index.items():
            if len(track_ids) == 1 and isrc in compact_stem:
                candidates.update(track_ids)

        # UPC alone identifies a release, never a track.  Require the regular
        # title matcher to select one row inside that UPC with a clear gap.
        digit_runs = re.findall(r"\d{8,14}", compact_stem)
        matched_upcs = {
            upc
            for run in digit_runs
            for upc in upc_index
            if len(upc) >= 8 and upc in run
        }
        for upc in matched_upcs:
            scored = sorted(
                ((track, *filename_match_score(audio, track)) for track in upc_index[upc]),
                key=lambda item: (-item[1], item[0].stable_id),
            )
            if scored and scored[0][1] >= 90 and (
                len(scored) == 1 or scored[1][1] < scored[0][1] - 2
            ):
                candidates.add(scored[0][0].stable_id)

        mapped_candidates = candidates.intersection(already_mapped)
        if mapped_candidates:
            ambiguous_files += int(bool(candidates.difference(already_mapped)))
            rejected_files += int(not candidates.difference(already_mapped))
            continue
        candidates = {
            track_id
            for track_id in candidates
            if track_id in by_id
            and not incompatible_version(audio.name, by_id[track_id].title)
        }
        if len(candidates) != 1:
            ambiguous_files += int(len(candidates) > 1)
            rejected_files += int(not candidates)
            continue
        claims_by_track[next(iter(candidates))].append(audio)

    additions: dict[tuple[str, str], list[DriveAudio]] = defaultdict(list)
    mapped = 0
    ambiguous_tracks = 0
    for track_id, audio_files in claims_by_track.items():
        unique = {audio.file_id: audio for audio in audio_files}
        if len(unique) != 1:
            ambiguous_tracks += 1
            continue
        track = by_id[track_id]
        additions[(track.upc, track.release)].append(next(iter(unique.values())))
        mapped += 1

    merged = list(release_audio)
    for (upc, release), files in sorted(additions.items(), key=lambda item: (item[0][0], normalize_text(item[0][1]))):
        merged.append(ReleaseAudio(upc=upc, release=release, folder_id="", files=tuple(files)))
    return merged, {
        "available": len(central_audio),
        "mapped": mapped,
        "ambiguousFiles": ambiguous_files,
        "ambiguousTracks": ambiguous_tracks,
        "rejected": rejected_files,
    }


def _central_exact_track_keys(track: Track) -> set[str]:
    """Return conservative complete-filename keys for an exact title signal."""

    title = _compact_identifier(track.title)
    if not title:
        return set()
    keys = {title}
    artists = {
        _compact_identifier(artist)
        for artist in track.artists
        if _compact_identifier(artist)
    }
    if track.artists:
        joined = _compact_identifier(" ".join(track.artists))
        if joined:
            artists.add(joined)
    for artist in artists:
        keys.update({f"{artist}{title}", f"{title}{artist}"})
    release = _compact_identifier(track.release)
    if release:
        keys.update({f"{release}{title}", f"{title}{release}"})
    return keys


def _central_composite_track_keys(track: Track) -> dict[str, set[str]]:
    """Build only complete deterministic keys, split by evidence rule."""

    title = _compact_identifier(track.title)
    artist_title: set[str] = set()
    release_title: set[str] = set()
    if title:
        artists = {
            _compact_identifier(artist)
            for artist in track.artists
            if _compact_identifier(artist)
        }
        joined = _compact_identifier(" ".join(track.artists))
        if joined:
            artists.add(joined)
        for artist in artists:
            artist_title.update({f"{artist}{title}", f"{title}{artist}"})
        release = _compact_identifier(track.release)
        if release:
            release_title.update({f"{release}{title}", f"{title}{release}"})
    return {
        "title": {title} if title else set(),
        "artist_title": artist_title,
        "release_title": release_title,
    }


def build_central_audio_pins(
    tracks: Sequence[Track],
    baseline_candidates: Sequence[Candidate],
    central_audio: Sequence[DriveAudio],
    *,
    excluded_file_ids: Iterable[str] = (),
) -> tuple[dict[int, CentralAudioPin], dict[str, Any]]:
    """Bind only new, conflict-free central files to exact source rows.

    Every available signal is exact: unique ISRC, UPC plus a complete title
    key, artist plus title, release plus title, and the established WAV-only
    globally unique title fallback. All detected signals must agree on one row;
    MP3 never receives the title-only fallback.

    Existing and ambiguous baseline candidates are never reconsidered. This
    makes a larger inventory monotone: it can fill a missing row, but cannot
    subtract or weaken an association that the previous snapshot established.
    """

    excluded = set(excluded_file_ids)
    tracks_by_row: dict[int, list[Track]] = defaultdict(list)
    candidates_by_row: dict[int, list[Candidate]] = defaultdict(list)
    isrc_rows: dict[str, set[int]] = defaultdict(set)
    upc_rows: dict[str, set[int]] = defaultdict(set)
    compact_title_rows: dict[str, set[int]] = defaultdict(set)
    artist_title_rows: dict[str, set[int]] = defaultdict(set)
    release_title_rows: dict[str, set[int]] = defaultdict(set)
    exact_keys_by_row: dict[int, dict[str, set[str]]] = {}
    for track in tracks:
        tracks_by_row[track.source_row].append(track)
        key_groups = _central_composite_track_keys(track)
        exact_keys_by_row[track.source_row] = key_groups
        compact_isrc = _compact_identifier(track.isrc)
        if compact_isrc:
            isrc_rows[compact_isrc].add(track.source_row)
        if track.upc:
            upc_rows[track.upc].add(track.source_row)
        for key in key_groups["title"]:
            compact_title_rows[key].add(track.source_row)
        for key in key_groups["artist_title"]:
            artist_title_rows[key].add(track.source_row)
        for key in key_groups["release_title"]:
            release_title_rows[key].add(track.source_row)
    for candidate in baseline_candidates:
        if candidate.track is not None:
            candidates_by_row[candidate.track.source_row].append(candidate)

    # Duplicate source rows are not identities and therefore cannot receive a
    # pin. The same rule applies to a row represented by two candidate records.
    eligible_rows = {
        row
        for row, source_tracks in tracks_by_row.items()
        if len(source_tracks) == 1 and len(candidates_by_row.get(row, ())) == 1
    }
    isrc_lengths = sorted({len(value) for value in isrc_rows})
    known_upcs = {
        upc
        for upc, rows in upc_rows.items()
        if len(upc) >= 8 and any(row in eligible_rows for row in rows)
    }
    upc_lengths = sorted({len(value) for value in known_upcs})
    claims_by_row: dict[int, list[tuple[DriveAudio, str]]] = defaultdict(list)
    ambiguous_files = 0
    rejected_files = 0
    conflicts = 0
    considered = 0
    unsupported_files = 0
    considered_by_format = {"wav": 0, "mp3": 0}
    for audio in central_audio:
        if audio.file_id in excluded:
            continue
        considered += 1
        descriptor = audio_source_descriptor(audio)
        if descriptor is None:
            rejected_files += 1
            unsupported_files += 1
            continue
        source_format, _source_mime_type = descriptor
        considered_by_format[source_format] += 1
        if is_suspicious_audio_name(audio.name):
            rejected_files += 1
            continue
        normalized_stem = _audio_stem(audio.name)
        compact_stem = _compact_identifier(normalized_stem)
        if not compact_stem:
            rejected_files += 1
            continue
        isrc_candidates: set[int] = set()
        matched_isrcs: set[str] = set()
        for length in isrc_lengths:
            for start in range(len(compact_stem) - length + 1):
                identifier = compact_stem[start : start + length]
                rows = isrc_rows.get(identifier)
                if rows:
                    matched_isrcs.add(identifier)
                    isrc_candidates.update(rows)

        matched_upcs: set[str] = set()
        for run in re.findall(r"\d{8,}", compact_stem):
            for length in upc_lengths:
                for start in range(len(run) - length + 1):
                    value = run[start : start + length]
                    if value in known_upcs:
                        matched_upcs.add(value)
        upc_title_candidates: set[int] = set()
        for upc in matched_upcs:
            without_upc = compact_stem.replace(upc, "")
            for row in upc_rows[upc]:
                if row not in eligible_rows:
                    continue
                key_groups = exact_keys_by_row.get(row, {})
                exact_after_upc = set().union(*key_groups.values()) if key_groups else set()
                if without_upc in exact_after_upc:
                    upc_title_candidates.add(row)

        identifier_variants = {compact_stem}
        for identifier in matched_isrcs.union(matched_upcs):
            identifier_variants.add(compact_stem.replace(identifier, ""))
        without_identifiers = compact_stem
        for identifier in sorted(
            matched_isrcs.union(matched_upcs), key=len, reverse=True
        ):
            without_identifiers = without_identifiers.replace(identifier, "")
        identifier_variants.add(without_identifiers)
        artist_title_candidates: set[int] = set()
        release_title_candidates: set[int] = set()
        title_candidates: set[int] = set()
        for variant in identifier_variants:
            artist_title_candidates.update(artist_title_rows.get(variant, set()))
            release_title_candidates.update(release_title_rows.get(variant, set()))
            # Preserve the previous exact-title WAV recovery.  Central MP3s
            # are intentionally narrower: identifier or full composite only.
            if source_format == "wav":
                rows = compact_title_rows.get(variant, set())
                if len(rows) == 1:
                    title_candidates.update(rows)

        rule_candidates = {
            "central_unique_isrc": isrc_candidates,
            "central_unique_upc_title": upc_title_candidates,
            "central_unique_artist_title": artist_title_candidates,
            "central_unique_release_title": release_title_candidates,
            "central_globally_unique_exact_title": title_candidates,
        }
        # Evidence tiers must agree on the same row.  A stronger identifier is
        # not allowed to hide a contradictory exact composite claim.
        candidates = set().union(*rule_candidates.values())
        if len(candidates) != 1:
            ambiguous_files += int(len(candidates) > 1)
            rejected_files += int(not candidates)
            conflicts += int(len(candidates) > 1)
            continue
        row = next(iter(candidates))
        if row not in eligible_rows:
            rejected_files += 1
            continue
        track = tracks_by_row[row][0]
        if incompatible_version(audio.name, track.title):
            rejected_files += 1
            continue
        match_kind = next(
            kind
            for kind in CENTRAL_AUDIO_PIN_RULES
            if row in rule_candidates[kind]
        )
        baseline = candidates_by_row[row][0]
        # Never turn an existing ambiguity into apparent uniqueness by adding
        # or subtracting files from its candidate set.
        if (
            baseline.audio is not None
            or "audio_match_missing" not in baseline.reasons
            or "audio_match_ambiguous" in baseline.reasons
            or "audio_file_claimed_by_multiple_tracks" in baseline.reasons
        ):
            rejected_files += 1
            continue
        claims_by_row[row].append((audio, match_kind))

    pins: dict[int, CentralAudioPin] = {}
    ambiguous_tracks = 0
    pinned_by_format = {"wav": 0, "mp3": 0}
    pinned_by_rule = {kind: 0 for kind in CENTRAL_AUDIO_PIN_RULES}
    for row, claims in claims_by_row.items():
        unique_files = {audio.file_id: (audio, kind) for audio, kind in claims}
        if len(unique_files) != 1:
            ambiguous_tracks += 1
            continue
        audio, kind = next(iter(unique_files.values()))
        track = tracks_by_row[row][0]
        pins[row] = CentralAudioPin(
            source_row=row,
            track_stable_id=track.stable_id,
            audio=audio,
            match_kind=kind,
        )
        source_format, _source_mime_type = audio_source_descriptor(audio) or ("", "")
        if source_format in pinned_by_format:
            pinned_by_format[source_format] += 1
        pinned_by_rule[kind] += 1
    return pins, {
        "available": len(central_audio),
        "considered": considered,
        "pinned": len(pins),
        "ambiguousFiles": ambiguous_files,
        "ambiguousTracks": ambiguous_tracks,
        "conflicts": conflicts,
        "rejected": rejected_files,
        "unsupported": unsupported_files,
        "consideredWav": considered_by_format["wav"],
        "consideredMp3": considered_by_format["mp3"],
        "pinnedWav": pinned_by_format["wav"],
        "pinnedMp3": pinned_by_format["mp3"],
        "pinnedByRule": dict(sorted(pinned_by_rule.items())),
    }


def merge_central_cover_mappings(
    covers: Mapping[str, Cover],
    tracks: Sequence[Track],
    artwork_rows: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Cover], dict[str, int]]:
    """Fill missing covers from one uniquely highest-ranked central image."""

    merged = dict(covers)
    release_upcs: dict[str, set[str]] = defaultdict(set)
    known_upcs = {track.upc for track in tracks if track.upc}
    for track in tracks:
        if track.upc and track.normalized_release:
            release_upcs[track.normalized_release].add(track.upc)
    candidates: dict[str, dict[str, tuple[int, str]]] = defaultdict(dict)
    for row in artwork_rows:
        file_id = extract_drive_id(row_value(row, "id", "file_id", "webViewLink"), folder=False)
        name = clean(row_value(row, "name", "file_name"))
        if not file_id or not name:
            continue
        stem = normalize_text(Path(name).stem)
        digits = re.sub(r"\D", "", name)
        matched = {upc for upc in known_upcs if len(upc) >= 8 and upc in digits}
        if not matched:
            matched = set(release_upcs.get(stem, set()))
        if len(matched) != 1:
            continue
        upc = next(iter(matched))
        if upc in merged:
            continue
        path = normalize_text(row_value(row, "path", "folder_path", "folderPath"))
        priority = 2 if "artwork les bon" in path else 1
        candidates[upc][file_id] = (priority, name)

    mapped = 0
    ambiguous = 0
    for upc, by_file in candidates.items():
        ranked = sorted(
            ((priority, normalize_text(name), file_id, name) for file_id, (priority, name) in by_file.items()),
            key=lambda item: (-item[0], item[1], item[2]),
        )
        best_priority = ranked[0][0]
        best = [row for row in ranked if row[0] == best_priority]
        if len(best) != 1:
            ambiguous += 1
            continue
        _priority, _normalized, file_id, name = best[0]
        merged[upc] = Cover(upc=upc, file_id=file_id, quality=name, is_square=False)
        mapped += 1
    return merged, {"available": len(artwork_rows), "mapped": mapped, "ambiguous": ambiguous}


def drive_inventory_is_complete(path: Path | None) -> bool:
    if path is None or path.suffix.casefold() != ".json" or not path.is_file():
        return False
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return False
    return isinstance(payload, dict) and payload.get("complete") is True


def source_inventories_are_complete(
    drive_inventory: Path | None,
    central_inventory: Path | None,
    central_complete: bool,
) -> bool:
    return drive_inventory_is_complete(drive_inventory) and (
        central_inventory is None or central_complete
    )


def merge_drive_cover_fallbacks(
    covers: Mapping[str, Cover],
    release_audio: Sequence[ReleaseAudio],
    inventory_children: Mapping[str, list[dict[str, Any]]],
) -> dict[str, Cover]:
    """Fill a missing workbook cover from one deterministic owned release image.

    The fallback never leaves the matched release folder. A unique image is
    sufficient; with several images, a clearly better cover/front/artwork name
    is required. Ambiguous artwork remains absent so the publication lane can
    exclude it instead of inventing a choice.
    """

    merged = dict(covers)
    groups_by_upc: dict[str, list[ReleaseAudio]] = defaultdict(list)
    for group in release_audio:
        if group.upc and group.folder_id:
            groups_by_upc[group.upc].append(group)

    def release_images(folder_id: str) -> list[dict[str, Any]]:
        folders = [folder_id]
        seen_folders: set[str] = set()
        images: dict[str, dict[str, Any]] = {}
        while folders:
            parent = folders.pop()
            if parent in seen_folders:
                continue
            seen_folders.add(parent)
            for row in inventory_children.get(parent, []):
                file_id = extract_drive_id(row_value(row, "id", "file_id", "webViewLink"), folder=False)
                if not file_id:
                    continue
                mime_type = clean(row_value(row, "mimeType", "mime_type")).casefold()
                name = clean(row_value(row, "name", "file_name"))
                if mime_type == DRIVE_FOLDER_MIME:
                    folders.append(file_id)
                    continue
                if mime_type.startswith("image/") or Path(name).suffix.casefold() in {
                    ".jpg",
                    ".jpeg",
                    ".png",
                    ".webp",
                }:
                    images.setdefault(file_id, {"id": file_id, "name": name})
        return list(images.values())

    def image_score(name: str, upc: str, release: str) -> int:
        normalized = normalize_text(Path(name).stem)
        tokens = set(normalized.split())
        if tokens.intersection({"back", "booklet", "banner", "canvas", "youtube", "story"}):
            return -1_000
        score = 0
        if normalized in {"cover", "artwork", "front", "folder"}:
            score += 200
        elif tokens.intersection({"cover", "artwork", "front"}):
            score += 120
        if upc and upc in re.sub(r"\D", "", name):
            score += 60
        normalized_release = normalize_text(release)
        if normalized_release and normalized_release in normalized:
            score += 40
        return score

    for upc, groups in groups_by_upc.items():
        if upc in merged:
            continue
        ranked_by_id: dict[str, tuple[int, str]] = {}
        for group in groups:
            for row in release_images(group.folder_id):
                score = image_score(str(row["name"]), upc, group.release)
                if score < 0:
                    continue
                previous = ranked_by_id.get(str(row["id"]))
                if previous is None or score > previous[0]:
                    ranked_by_id[str(row["id"])] = (score, str(row["name"]))
        ranked = sorted(
            ((score, name, file_id) for file_id, (score, name) in ranked_by_id.items()),
            key=lambda item: (-item[0], normalize_text(item[1]), item[2]),
        )
        if not ranked:
            continue
        if len(ranked) > 1 and (ranked[0][0] < 80 or ranked[0][0] < ranked[1][0] + 20):
            continue
        _score, name, file_id = ranked[0]
        merged[upc] = Cover(upc=upc, file_id=file_id, quality=name, is_square=False)
    return merged


def drive_audio_from_row(row: Mapping[str, Any], *, folder_id: str = "", folder_path: str = "") -> DriveAudio:
    size_raw = row_value(row, "size", "size_bytes")
    try:
        size = int(size_raw) if clean(size_raw) else None
    except (TypeError, ValueError):
        size = None
    return DriveAudio(
        file_id=extract_drive_id(row_value(row, "id", "file_id", "webViewLink"), folder=False),
        name=clean(row_value(row, "name", "file_name")),
        folder_id=folder_id,
        folder_path=folder_path,
        mime_type=clean(row_value(row, "mimeType", "mime_type")),
        size_bytes=size,
        modified_time=clean(row_value(row, "modifiedTime", "modified_time")),
    )


class GoogleDriveClient:
    def __init__(self, access_token: str = "") -> None:
        self.access_token = access_token

    def _request(self, url: str, *, method: str = "GET", headers: Mapping[str, str] | None = None, timeout: int = 60) -> Any:
        merged = {"User-Agent": "SymbiomeCatalogAudit/1.0", **(dict(headers or {}))}
        if self.access_token:
            merged["Authorization"] = f"Bearer {self.access_token}"
        request = urllib.request.Request(url, method=method, headers=merged)
        return urllib.request.urlopen(request, timeout=timeout)

    def list_children(self, folder_id: str) -> list[dict[str, Any]]:
        if not self.access_token:
            raise RuntimeError("Drive folder discovery requires GOOGLE_DRIVE_ACCESS_TOKEN or a private --drive-inventory export.")
        files: list[dict[str, Any]] = []
        page_token = ""
        while True:
            query = {
                "q": f"'{folder_id}' in parents and trashed = false",
                "fields": "nextPageToken,files(id,name,mimeType,size,md5Checksum,modifiedTime,parents,webViewLink)",
                "pageSize": "1000",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            }
            if page_token:
                query["pageToken"] = page_token
            url = f"https://www.googleapis.com/drive/v3/files?{urllib.parse.urlencode(query)}"
            with self._request(url) as response:
                payload = json.load(response)
            files.extend(payload.get("files", []))
            page_token = clean(payload.get("nextPageToken"))
            if not page_token:
                return files

    def discover_wavs(
        self,
        release_folder_id: str,
        *,
        inventory_children: Mapping[str, list[dict[str, Any]]] | None = None,
        max_depth: int = 6,
    ) -> list[DriveAudio]:
        inventory_children = inventory_children or {}
        queue: list[tuple[str, tuple[str, ...], int]] = [(release_folder_id, (), 0)]
        seen_folders: set[str] = set()
        audio: dict[str, DriveAudio] = {}
        while queue:
            folder_id, path_parts, depth = queue.pop(0)
            if folder_id in seen_folders or depth > max_depth:
                continue
            seen_folders.add(folder_id)
            children = inventory_children.get(folder_id)
            if children is None:
                children = self.list_children(folder_id)
            for child in children:
                child_id = extract_drive_id(row_value(child, "id", "webViewLink"), folder=False)
                if not child_id:
                    continue
                name = clean(row_value(child, "name"))
                mime_type = clean(row_value(child, "mimeType", "mime_type"))
                if mime_type == DRIVE_FOLDER_MIME:
                    queue.append((child_id, (*path_parts, name), depth + 1))
                    continue
                if name.casefold().endswith(".wav") or mime_type.casefold() in WAV_MIME_TYPES:
                    # Direct WAVs are valid. AUDIO_FOLDER_NAMES is used as a
                    # discovery hint and recorded in path; it is not a filter.
                    path_text = "/".join(path_parts)
                    audio.setdefault(child_id, drive_audio_from_row(child, folder_id=folder_id, folder_path=path_text))
        return list(audio.values())

    def media_url(self, file_id: str) -> str:
        if self.access_token:
            return f"https://www.googleapis.com/drive/v3/files/{urllib.parse.quote(file_id)}?alt=media&supportsAllDrives=true"
        return f"https://drive.google.com/uc?export=download&id={urllib.parse.quote(file_id)}"

    def read_range(self, file_id: str, *, end: int = 1_048_575) -> tuple[bytes, dict[str, str]]:
        with self._request(self.media_url(file_id), headers={"Range": f"bytes=0-{end}"}, timeout=120) as response:
            payload = response.read(end + 1)
            headers = {key.casefold(): value for key, value in response.headers.items()}
        reject_html(payload, headers)
        return payload, headers

    def hash_full_file(self, file_id: str) -> tuple[str, bytes, dict[str, str], int]:
        temporary_path = ""
        digest = hashlib.sha256()
        prefix = bytearray()
        size = 0
        try:
            with self._request(self.media_url(file_id), timeout=600) as response:
                headers = {key.casefold(): value for key, value in response.headers.items()}
                with tempfile.NamedTemporaryFile(prefix="symbiome-wav-", suffix=".part", delete=False) as temporary:
                    temporary_path = temporary.name
                    while True:
                        chunk = response.read(1024 * 1024)
                        if not chunk:
                            break
                        temporary.write(chunk)
                        digest.update(chunk)
                        size += len(chunk)
                        if len(prefix) < 1024 * 1024:
                            prefix.extend(chunk[: 1024 * 1024 - len(prefix)])
            reject_html(bytes(prefix), headers)
            return digest.hexdigest(), bytes(prefix), headers, size
        finally:
            if temporary_path:
                try:
                    os.unlink(temporary_path)
                except FileNotFoundError:
                    pass


def reject_html(payload: bytes, headers: Mapping[str, str]) -> None:
    content_type = clean(headers.get("content-type")).casefold()
    if "text/html" in content_type or payload.lstrip().lower().startswith((b"<!doctype html", b"<html")):
        raise RuntimeError("Drive returned an HTML permission/confirmation page instead of audio.")


def parse_wav_prefix(payload: bytes) -> WavInfo:
    if len(payload) < 12:
        raise ValueError("WAV header is truncated")
    container_bytes, riff_size, wave_marker = struct.unpack_from("<4sI4s", payload, 0)
    if container_bytes not in {b"RIFF", b"RF64"} or wave_marker != b"WAVE":
        raise ValueError("not a RIFF/RF64 WAVE file")
    offset = 12
    fmt: tuple[int, int, int, int, int] | None = None
    data_size: int | None = None
    rf64_data_size: int | None = None
    while offset + 8 <= len(payload):
        chunk_id, chunk_size = struct.unpack_from("<4sI", payload, offset)
        chunk_start = offset + 8
        chunk_end = chunk_start + chunk_size
        if chunk_id == b"ds64" and chunk_start + 24 <= len(payload):
            _riff_size_64, rf64_data_size, _sample_count = struct.unpack_from("<QQQ", payload, chunk_start)
        elif chunk_id == b"fmt " and chunk_start + 16 <= len(payload):
            audio_format, channels, sample_rate, byte_rate, _block_align, bits = struct.unpack_from("<HHIIHH", payload, chunk_start)
            fmt = (audio_format, channels, sample_rate, byte_rate, bits)
        elif chunk_id == b"data":
            data_size = rf64_data_size if chunk_size == 0xFFFFFFFF and rf64_data_size is not None else chunk_size
            break
        if chunk_end > len(payload):
            break
        offset = chunk_end + (chunk_size % 2)
    if fmt is None:
        raise ValueError("WAV fmt chunk was not found in the inspected range")
    if data_size is None:
        raise ValueError("WAV data chunk was not found in the inspected range")
    audio_format, channels, sample_rate, byte_rate, bits = fmt
    if byte_rate <= 0:
        raise ValueError("WAV byte rate is invalid")
    codecs = {1: "PCM", 3: "IEEE_FLOAT", 6: "ALAW", 7: "MULAW", 0xFFFE: "EXTENSIBLE"}
    return WavInfo(
        container=container_bytes.decode("ascii"),
        codec=codecs.get(audio_format, f"WAVE_FORMAT_{audio_format}"),
        channels=channels,
        sample_rate=sample_rate,
        bits_per_sample=bits,
        byte_rate=byte_rate,
        data_size=data_size,
        duration_seconds=data_size / byte_rate,
    )


def filename_variants(audio: DriveAudio, track: Track) -> set[str]:
    stem = re.sub(r"\.wav$", "", audio.name, flags=re.IGNORECASE)
    stem = re.sub(r"^\s*\d{1,3}\s*[._-]+\s*", "", stem)
    normalized = normalize_text(stem)
    variants = {normalized}
    removable = [normalize_text(track.release), *(normalize_text(artist) for artist in track.artists)]
    for part in removable:
        if not part:
            continue
        variants.update(
            {
                value.strip()
                for value in (
                    normalized.removeprefix(f"{part} "),
                    normalized.removesuffix(f" {part}"),
                    normalized.replace(f" {part} ", " "),
                )
                if value.strip()
            }
        )
    return variants


def filename_match_score(audio: DriveAudio, track: Track) -> tuple[float, str]:
    target = track.normalized_title
    if not target:
        return 0, "missing_track_title"
    best = 0.0
    kind = "no_match"
    for variant in filename_variants(audio, track):
        if variant == target:
            return 100.0, "exact_title"
        if variant.endswith(f" {target}") or variant.startswith(f"{target} "):
            if best < 96:
                best, kind = 96.0, "title_boundary"
        elif re.search(rf"(?:^| ){re.escape(target)}(?: |$)", variant):
            if best < 93:
                best, kind = 93.0, "title_contained"
        ratio = SequenceMatcher(None, variant, target, autojunk=False).ratio() * 100
        if ratio > best:
            best, kind = ratio, "fuzzy_title"
    return round(best, 2), kind


def incompatible_version(audio_name: str, title: str) -> bool:
    audio_tokens = set(normalize_text(audio_name).split()).intersection(VERSION_TOKENS)
    title_tokens = set(normalize_text(title).split()).intersection(VERSION_TOKENS)
    return bool(audio_tokens.symmetric_difference(title_tokens))


def is_suspicious_audio_name(name: str) -> bool:
    normalized = normalize_text(name)
    return any(token in normalized for token in PARASITE_TOKENS)


def orchard_indices(rows: Sequence[OrchardTrack]) -> dict[str, dict[tuple[str, ...], list[OrchardTrack]]]:
    maps: dict[str, dict[tuple[str, ...], list[OrchardTrack]]] = {
        "strict": defaultdict(list),
        "no_artist": defaultdict(list),
        "upc_title": defaultdict(list),
    }
    for row in rows:
        upc = row.upc
        release = normalize_text(row.release)
        title = normalize_text(row.title)
        artist = normalize_text(row.artist)
        maps["strict"][(upc, release, title, artist)].append(row)
        maps["no_artist"][(upc, release, title)].append(row)
        maps["upc_title"][(upc, title)].append(row)
    return maps


def match_orchard(track: Track, maps: Mapping[str, Mapping[tuple[str, ...], list[OrchardTrack]]]) -> tuple[str, OrchardTrack | None]:
    artists = {normalize_text(artist) for artist in track.artists if artist}
    if track.artists:
        artists.add(normalize_text(" ".join(track.artists)))
    strict: dict[str, OrchardTrack] = {}
    for artist in artists:
        for row in maps["strict"].get((track.upc, track.normalized_release, track.normalized_title, artist), []):
            strict[row.spotify_id or f"no-id-{id(row)}"] = row
    if len(strict) == 1:
        return "exact", next(iter(strict.values()))
    if len(strict) > 1:
        return "ambiguous_exact", None
    if maps["no_artist"].get((track.upc, track.normalized_release, track.normalized_title)):
        return "artist_mismatch", None
    if maps["upc_title"].get((track.upc, track.normalized_title)):
        return "release_mismatch", None
    return "missing", None


def match_release_audio(track: Track, release_audio: Sequence[ReleaseAudio]) -> list[DriveAudio]:
    exact_upc = [group for group in release_audio if track.upc and group.upc == track.upc]
    exact_both = [group for group in exact_upc if normalize_text(group.release) == track.normalized_release]
    selected = exact_both or exact_upc
    if not selected:
        selected = [group for group in release_audio if normalize_text(group.release) == track.normalized_release]
    files: dict[str, DriveAudio] = {}
    for group in selected:
        for audio in group.files:
            files.setdefault(audio.file_id, audio)
    return list(files.values())


def candidate_fingerprint(track: Track | None, audio: DriveAudio | None, spotify_id: str, cover: Cover | None) -> str:
    raw = json.dumps(
        {
            "track": dataclasses.asdict(track) if track else None,
            "audio": dataclasses.asdict(audio) if audio else None,
            "spotify": spotify_id,
            "cover": dataclasses.asdict(cover) if cover else None,
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def build_candidates(
    tracks: Sequence[Track],
    release_audio: Sequence[ReleaseAudio],
    covers: Mapping[str, Cover],
    orchard_rows: Sequence[OrchardTrack],
) -> list[Candidate]:
    orchard_maps = orchard_indices(orchard_rows)
    releases_by_upc: dict[str, list[ReleaseAudio]] = defaultdict(list)
    releases_by_name: dict[str, list[ReleaseAudio]] = defaultdict(list)
    releases_by_both: dict[tuple[str, str], list[ReleaseAudio]] = defaultdict(list)
    for group in release_audio:
        normalized_release = normalize_text(group.release)
        if group.upc:
            releases_by_upc[group.upc].append(group)
            releases_by_both[(group.upc, normalized_release)].append(group)
        if normalized_release:
            releases_by_name[normalized_release].append(group)
    preliminary: list[Candidate] = []
    claimed_audio: dict[str, list[tuple[int, float]]] = defaultdict(list)
    all_audio: dict[str, DriveAudio] = {}

    for track in tracks:
        groups = releases_by_both.get((track.upc, track.normalized_release))
        if not groups:
            groups = releases_by_upc.get(track.upc) if track.upc else None
        if not groups:
            groups = releases_by_name.get(track.normalized_release, [])
        available_by_id: dict[str, DriveAudio] = {}
        for group in groups:
            for grouped_audio in group.files:
                available_by_id.setdefault(grouped_audio.file_id, grouped_audio)
        available = list(available_by_id.values())
        all_audio.update({audio.file_id: audio for audio in available})
        scores = sorted(
            ((audio, *filename_match_score(audio, track)) for audio in available),
            key=lambda item: (-item[1], normalize_text(item[0].name), item[0].file_id),
        )
        audio: DriveAudio | None = None
        score: float | None = None
        match_kind = "missing"
        reasons: list[str] = []
        if scores and scores[0][1] >= 90:
            top = scores[0]
            tied = len(scores) > 1 and scores[1][1] >= top[1] - 2
            if tied:
                match_kind = "ambiguous"
                reasons.append("audio_match_ambiguous")
            else:
                audio, score, match_kind = top
        else:
            reasons.append("audio_match_missing")

        spotify_match_kind, orchard = match_orchard(track, orchard_maps) if orchard_rows else ("source_missing", None)
        spotify_id = orchard.spotify_id if orchard else ""
        spotify_duration = orchard.spotify_duration_seconds if orchard else None
        cover = covers.get(track.upc)

        if audio and incompatible_version(audio.name, track.title):
            reasons.append("audio_version_mismatch")
        if audio and is_suspicious_audio_name(audio.name):
            reasons.append("suspicious_audio_filename")
        if track.duration_seconds is None:
            reasons.append("expected_duration_missing")
        elif track.duration_seconds < 30:
            reasons.append("expected_duration_under_30s")
        if spotify_match_kind != "exact":
            reasons.append(f"spotify_{spotify_match_kind}")
        elif not spotify_id:
            reasons.append("spotify_id_missing")
        if spotify_duration is None:
            reasons.append("spotify_duration_missing")
        if cover is None:
            reasons.append("cover_missing")
        elif not cover.is_square:
            reasons.append("cover_not_verified_square")
        reasons.append("audio_inspection_pending")
        if audio:
            reasons.append("sha256_pending")

        quarantine_reasons = {
            "audio_match_missing",
            "audio_match_ambiguous",
            "audio_version_mismatch",
            "suspicious_audio_filename",
            "expected_duration_under_30s",
        }
        status = "quarantine" if quarantine_reasons.intersection(reasons) else "review"
        candidate_id = hashlib.sha256(f"{track.stable_id}|{audio.stable_id if audio else 'missing'}".encode()).hexdigest()[:28]
        candidate = Candidate(
            candidate_id=candidate_id,
            track=track,
            audio=audio,
            audio_match_score=score,
            audio_match_kind=match_kind,
            spotify_id=spotify_id,
            spotify_duration_seconds=spotify_duration,
            spotify_match_kind=spotify_match_kind,
            cover=cover,
            status=status,
            reasons=sorted(set(reasons)),
            fingerprint=candidate_fingerprint(track, audio, spotify_id, cover),
        )
        preliminary.append(candidate)
        if audio and score is not None:
            claimed_audio[audio.file_id].append((len(preliminary) - 1, score))

    # A Drive file cannot silently satisfy several catalogue rows. Keep only a
    # clearly better claim; ties are quarantined for manual review.
    for claims in claimed_audio.values():
        if len(claims) < 2:
            continue
        claims.sort(key=lambda item: item[1], reverse=True)
        winner_is_clear = claims[0][1] >= claims[1][1] + 8
        affected = claims[1:] if winner_is_clear else claims
        for index, _score in affected:
            candidate = preliminary[index]
            candidate.status = "quarantine"
            candidate.reasons = sorted(set([*candidate.reasons, "audio_file_claimed_by_multiple_tracks"]))

    used_audio_ids = {candidate.audio.file_id for candidate in preliminary if candidate.audio}
    for audio_id, audio in sorted(all_audio.items(), key=lambda item: (normalize_text(item[1].name), item[0])):
        if audio_id in used_audio_ids:
            continue
        candidate_id = hashlib.sha256(f"orphan|{audio.stable_id}".encode()).hexdigest()[:28]
        preliminary.append(
            Candidate(
                candidate_id=candidate_id,
                track=None,
                audio=audio,
                audio_match_score=None,
                audio_match_kind="orphan",
                spotify_id="",
                spotify_duration_seconds=None,
                spotify_match_kind="missing",
                cover=None,
                status="quarantine",
                reasons=["unmatched_audio_file"],
                fingerprint=candidate_fingerprint(None, audio, "", None),
            )
        )
    return preliminary


def apply_central_audio_pins(
    candidates: Sequence[Candidate],
    pins: Mapping[int, CentralAudioPin],
) -> list[Candidate]:
    """Apply exact row pins without changing any pre-existing association."""

    result: list[Candidate] = []
    for candidate in candidates:
        track = candidate.track
        pin = pins.get(track.source_row) if track is not None else None
        if (
            pin is None
            or candidate.audio is not None
            or "audio_match_missing" not in candidate.reasons
            or "audio_match_ambiguous" in candidate.reasons
            or "audio_file_claimed_by_multiple_tracks" in candidate.reasons
            or track is None
            or pin.track_stable_id != track.stable_id
        ):
            result.append(candidate)
            continue
        reasons = sorted(
            set(
                [
                    *(reason for reason in candidate.reasons if reason != "audio_match_missing"),
                    "sha256_pending",
                ]
            )
        )
        quarantine_reasons = {
            "audio_match_missing",
            "audio_match_ambiguous",
            "audio_version_mismatch",
            "suspicious_audio_filename",
            "expected_duration_under_30s",
        }
        status = "quarantine" if quarantine_reasons.intersection(reasons) else "review"
        candidate_id = hashlib.sha256(
            f"{track.stable_id}|{pin.audio.stable_id}".encode()
        ).hexdigest()[:28]
        result.append(
            dataclasses.replace(
                candidate,
                candidate_id=candidate_id,
                audio=pin.audio,
                audio_match_score=100.0,
                audio_match_kind=pin.match_kind,
                status=status,
                reasons=reasons,
                fingerprint=candidate_fingerprint(
                    track, pin.audio, candidate.spotify_id, candidate.cover
                ),
            )
        )
    return result


def open_state(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=DELETE")
    connection.execute("PRAGMA foreign_keys=ON")
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS candidates (
          candidate_id TEXT PRIMARY KEY,
          fingerprint TEXT NOT NULL,
          payload_json TEXT NOT NULL,
          status TEXT NOT NULL CHECK (status IN ('exact', 'review', 'quarantine')),
          reasons_json TEXT NOT NULL,
          inspection_status TEXT NOT NULL DEFAULT 'pending',
          inspection_mode TEXT,
          content_type TEXT,
          content_length INTEGER,
          wav_json TEXT,
          sha256 TEXT,
          error TEXT,
          updated_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS candidates_status_idx ON candidates(status, inspection_status);
        """
    )
    return connection


def candidate_payload(candidate: Candidate) -> dict[str, Any]:
    audio_payload = dataclasses.asdict(candidate.audio) if candidate.audio else None
    if audio_payload is not None:
        descriptor = audio_source_descriptor(audio_payload)
        if descriptor is not None:
            audio_payload["source_format"], audio_payload["source_mime_type"] = descriptor
    return {
        "candidate_id": candidate.candidate_id,
        "track": dataclasses.asdict(candidate.track) if candidate.track else None,
        "audio": audio_payload,
        "audio_match_score": candidate.audio_match_score,
        "audio_match_kind": candidate.audio_match_kind,
        "spotify_id": candidate.spotify_id,
        "spotify_duration_seconds": candidate.spotify_duration_seconds,
        "spotify_match_kind": candidate.spotify_match_kind,
        "cover": dataclasses.asdict(candidate.cover) if candidate.cover else None,
    }


def direct_publication_eligible(record: Mapping[str, Any]) -> bool:
    """Return whether Sheet + Drive provide one deterministic publishable pair.

    This owner-authoritative lane deliberately ignores Spotify and historical
    inspection state.  It still refuses missing or multiply claimed audio,
    missing artwork and any non-deterministic filename association.  The
    publication worker performs the one full WAV download, checksum and decode
    before it can stage or promote the row.
    """

    track = record.get("track")
    audio = record.get("audio")
    cover = record.get("cover")
    reasons = record.get("reasons")
    score = record.get("audio_match_score")
    match_kind = clean(record.get("audio_match_kind"))
    if not isinstance(track, Mapping) or not isinstance(audio, Mapping) or not isinstance(cover, Mapping):
        return False
    if not isinstance(reasons, list) or any(reason in DIRECT_PUBLICATION_BLOCKERS for reason in reasons):
        return False
    if (
        isinstance(score, bool)
        or not isinstance(score, (int, float))
        or not math.isfinite(float(score))
        or float(score) < 90
        or match_kind in {"", "missing", "ambiguous", "orphan"}
    ):
        return False
    descriptor = audio_source_descriptor(audio)
    if descriptor is None:
        return False
    source_format, _source_mime_type = descriptor
    if source_format == "mp3" and match_kind not in STRICT_CENTRAL_MP3_MATCH_KINDS:
        return False
    source_row = track.get("source_row")
    if isinstance(source_row, bool) or not isinstance(source_row, int) or source_row < 1:
        return False
    return bool(
        DRIVE_ID_PATTERN.fullmatch(clean(audio.get("file_id")))
        and clean(audio.get("name"))
        and DRIVE_ID_PATTERN.fullmatch(clean(cover.get("file_id")))
    )


def upsert_candidates(connection: sqlite3.Connection, candidates: Sequence[Candidate], *, force: bool = False) -> None:
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    for candidate in candidates:
        existing = connection.execute(
            "SELECT fingerprint, payload_json, inspection_status, inspection_mode, content_type, content_length, wav_json, sha256, error, updated_at FROM candidates WHERE candidate_id = ?",
            (candidate.candidate_id,),
        ).fetchone()
        preserve_inspection = bool(existing and existing["fingerprint"] == candidate.fingerprint and not force)
        payload = candidate_payload(candidate)
        status = candidate.status
        reasons = list(candidate.reasons)
        if preserve_inspection:
            existing_payload = json.loads(existing["payload_json"])
            verified_spotify_duration = existing_payload.get("spotify_duration_seconds")
            if payload.get("spotify_duration_seconds") is None and isinstance(verified_spotify_duration, (int, float)):
                payload["spotify_duration_seconds"] = verified_spotify_duration
                reasons = [reason for reason in reasons if reason != "spotify_duration_missing"]
            wav_payload = json.loads(existing["wav_json"]) if existing["wav_json"] else None
            wav = WavInfo(**wav_payload) if wav_payload else None
            status, reasons = classify_after_inspection(
                payload,
                reasons,
                wav,
                clean(existing["sha256"]),
                clean(existing["error"]),
            )
        values = {
            "inspection_status": existing["inspection_status"] if preserve_inspection else "pending",
            "inspection_mode": existing["inspection_mode"] if preserve_inspection else None,
            "content_type": existing["content_type"] if preserve_inspection else None,
            "content_length": existing["content_length"] if preserve_inspection else None,
            "wav_json": existing["wav_json"] if preserve_inspection else None,
            "sha256": existing["sha256"] if preserve_inspection else None,
            "error": existing["error"] if preserve_inspection else None,
            "updated_at": existing["updated_at"] if preserve_inspection else now,
        }
        connection.execute(
            """
            INSERT INTO candidates (
              candidate_id, fingerprint, payload_json, status, reasons_json,
              inspection_status, inspection_mode, content_type, content_length,
              wav_json, sha256, error, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(candidate_id) DO UPDATE SET
              fingerprint=excluded.fingerprint,
              payload_json=excluded.payload_json,
              status=excluded.status,
              reasons_json=excluded.reasons_json,
              inspection_status=excluded.inspection_status,
              inspection_mode=excluded.inspection_mode,
              content_type=excluded.content_type,
              content_length=excluded.content_length,
              wav_json=excluded.wav_json,
              sha256=excluded.sha256,
              error=excluded.error,
              updated_at=excluded.updated_at
            """,
            (
                candidate.candidate_id,
                candidate.fingerprint,
                json.dumps(payload, ensure_ascii=False, sort_keys=True),
                status,
                json.dumps(reasons, ensure_ascii=False),
                values["inspection_status"],
                values["inspection_mode"],
                values["content_type"],
                values["content_length"],
                values["wav_json"],
                values["sha256"],
                values["error"],
                values["updated_at"],
            ),
        )
    connection.commit()


def quarantine_stale_candidates(connection: sqlite3.Connection, current_candidate_ids: set[str]) -> int:
    """Hide rows absent from a complete refreshed workbook + Drive snapshot."""

    stale = 0
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    for row in connection.execute("SELECT candidate_id, status, reasons_json FROM candidates").fetchall():
        candidate_id = str(row["candidate_id"])
        if candidate_id in current_candidate_ids:
            continue
        reasons = set(json.loads(row["reasons_json"]))
        if row["status"] == "quarantine" and "source_snapshot_missing" in reasons:
            continue
        reasons.add("source_snapshot_missing")
        connection.execute(
            """UPDATE candidates SET status='quarantine', reasons_json=?, updated_at=?
            WHERE candidate_id=?""",
            (json.dumps(sorted(reasons), ensure_ascii=False), now, candidate_id),
        )
        stale += 1
    connection.commit()
    return stale


def content_length_from_headers(headers: Mapping[str, str], fallback: int) -> int:
    content_range = clean(headers.get("content-range"))
    match = re.search(r"/(\d+)$", content_range)
    if match:
        return int(match.group(1))
    try:
        return int(clean(headers.get("content-length")))
    except ValueError:
        return fallback


def classify_after_inspection(
    payload: Mapping[str, Any], reasons: Sequence[str], wav: WavInfo | None, sha256: str, error: str
) -> tuple[str, list[str]]:
    revised = {reason for reason in reasons if reason not in {"audio_inspection_pending", "sha256_pending"}}
    if error or wav is None:
        revised.add("audio_inspection_failed")
        return "quarantine", sorted(revised)

    track = payload.get("track") or {}
    expected = track.get("duration_seconds")
    spotify_duration = payload.get("spotify_duration_seconds")
    if wav.duration_seconds < 30:
        revised.add("wav_duration_under_30s")
    deltas: list[tuple[str, float, float]] = []
    for label, reference in (("catalogue", expected), ("spotify", spotify_duration)):
        if isinstance(reference, (int, float)) and reference > 0:
            delta = abs(wav.duration_seconds - float(reference))
            deltas.append((label, delta, float(reference)))
            if delta > 5 or delta / float(reference) > 0.02:
                revised.add(f"duration_{label}_blocked")
            elif delta > 2:
                revised.add(f"duration_{label}_review")
    if not sha256:
        revised.add("sha256_pending")

    quarantine_markers = {
        "audio_match_missing",
        "audio_match_ambiguous",
        "audio_version_mismatch",
        "suspicious_audio_filename",
        "expected_duration_under_30s",
        "audio_file_claimed_by_multiple_tracks",
        "unmatched_audio_file",
        "wav_duration_under_30s",
        "duration_catalogue_blocked",
        "duration_spotify_blocked",
        "audio_inspection_failed",
    }
    if quarantine_markers.intersection(revised):
        return "quarantine", sorted(revised)

    review_prefixes = ("spotify_", "cover_", "expected_duration_", "duration_", "sha256_")
    if any(reason.startswith(review_prefixes) for reason in revised):
        return "review", sorted(revised)
    return "exact", sorted(revised)


def apply_spotify_metadata(
    connection: sqlite3.Connection,
    spotify_metadata: Mapping[str, Mapping[str, Any]],
) -> dict[str, int]:
    """Attach verified Spotify metadata to staged candidates.

    The Orchard URI alone is never enough to publish.  This step records the
    Spotify duration and re-runs the same fail-closed classifier used after WAV
    inspection so duration evidence cannot be bolted on without revalidation.
    """

    counts = {"matched": 0, "updated": 0, "missing": 0}
    rows = connection.execute(
        "SELECT * FROM candidates WHERE json_extract(payload_json, '$.spotify_id') != ''"
    ).fetchall()
    for row in rows:
        payload = json.loads(row["payload_json"])
        spotify_id = clean(payload.get("spotify_id"))
        metadata = spotify_metadata.get(spotify_id)
        if not metadata:
            counts["missing"] += 1
            continue
        duration_ms = metadata.get("duration_ms")
        if not isinstance(duration_ms, (int, float)) or duration_ms <= 0:
            counts["missing"] += 1
            continue

        counts["matched"] += 1
        payload["spotify_duration_seconds"] = float(duration_ms) / 1000
        reasons = [
            reason
            for reason in json.loads(row["reasons_json"])
            if reason != "spotify_duration_missing"
        ]
        wav_payload = json.loads(row["wav_json"]) if row["wav_json"] else None
        wav = WavInfo(**wav_payload) if wav_payload else None
        status, revised = classify_after_inspection(
            payload,
            reasons,
            wav,
            clean(row["sha256"]),
            clean(row["error"]),
        )
        connection.execute(
            """
            UPDATE candidates
            SET payload_json=?, status=?, reasons_json=?, updated_at=?
            WHERE candidate_id=?
            """,
            (
                json.dumps(payload, ensure_ascii=False, sort_keys=True),
                status,
                json.dumps(revised, ensure_ascii=False),
                dt.datetime.now(dt.timezone.utc).isoformat(),
                row["candidate_id"],
            ),
        )
        counts["updated"] += 1
    connection.commit()
    return counts


def inspect_pending(
    connection: sqlite3.Connection,
    drive: GoogleDriveClient,
    *,
    mode: str,
    batch_size: int,
    force: bool,
    release_batch_size: int | None = None,
    maximum_seconds: int | None = None,
) -> dict[str, int]:
    if mode == "none":
        return {
            "selected": 0,
            "complete": 0,
            "failed": 0,
            "releases": 0,
            "remaining": 0,
            "remainingReleases": 0,
            "timeLimitReached": 0,
        }
    where = "json_extract(payload_json, '$.audio.file_id') IS NOT NULL"
    parameters: list[Any] = []
    if not force:
        where += " AND (inspection_status != 'complete' OR (? = 'full' AND COALESCE(sha256, '') = ''))"
        parameters.append(mode)
    pending_rows = connection.execute(
        f"""SELECT * FROM candidates WHERE {where}
        ORDER BY CASE inspection_status
          WHEN 'pending' THEN 0
          WHEN 'complete' THEN 1
          WHEN 'failed' THEN 2
          ELSE 1
        END,
        CASE WHEN inspection_status = 'failed' THEN updated_at ELSE '' END,
        candidate_id""",  # noqa: S608 - static clause
        parameters,
    ).fetchall()
    rows: list[sqlite3.Row] = []
    release_keys: set[str] = set()
    for row in pending_rows:
        payload = json.loads(row["payload_json"])
        track = payload.get("track") if isinstance(payload.get("track"), dict) else {}
        audio = payload.get("audio") if isinstance(payload.get("audio"), dict) else {}
        release_key = (
            normalize_upc(track.get("upc"))
            or clean(audio.get("folder_id"))
            or normalize_text(track.get("release"))
            or row["candidate_id"]
        )
        if release_key not in release_keys and release_batch_size is not None and len(release_keys) >= release_batch_size:
            continue
        release_keys.add(release_key)
        rows.append(row)
        if len(rows) >= batch_size:
            break
    counts = {
        "selected": 0,
        "complete": 0,
        "failed": 0,
        "releases": len(release_keys),
        "remaining": 0,
        "remainingReleases": 0,
        "timeLimitReached": 0,
    }
    started = time.monotonic()
    for row in rows:
        if maximum_seconds is not None and time.monotonic() - started >= maximum_seconds:
            counts["timeLimitReached"] = 1
            break
        counts["selected"] += 1
        payload = json.loads(row["payload_json"])
        file_id = payload["audio"]["file_id"]
        error = ""
        wav: WavInfo | None = None
        sha256 = ""
        headers: dict[str, str] = {}
        content_length = 0
        try:
            if mode == "full":
                sha256, prefix, headers, content_length = drive.hash_full_file(file_id)
            else:
                prefix, headers = drive.read_range(file_id)
                content_length = content_length_from_headers(headers, len(prefix))
            wav = parse_wav_prefix(prefix)
        except (OSError, ValueError, RuntimeError, urllib.error.URLError) as inspection_error:
            error = f"{type(inspection_error).__name__}: {inspection_error}"
        status, reasons = classify_after_inspection(payload, json.loads(row["reasons_json"]), wav, sha256, error)
        connection.execute(
            """
            UPDATE candidates SET status=?, reasons_json=?, inspection_status=?, inspection_mode=?,
              content_type=?, content_length=?, wav_json=?, sha256=?, error=?, updated_at=?
            WHERE candidate_id=?
            """,
            (
                status,
                json.dumps(reasons, ensure_ascii=False),
                "failed" if error else "complete",
                mode,
                clean(headers.get("content-type")),
                content_length or None,
                json.dumps(dataclasses.asdict(wav), ensure_ascii=False, sort_keys=True) if wav else None,
                sha256 or None,
                error or None,
                dt.datetime.now(dt.timezone.utc).isoformat(),
                row["candidate_id"],
            ),
        )
        connection.commit()  # resume safely after every file
        counts["failed" if error else "complete"] += 1
    remaining_rows = connection.execute(
        f"SELECT payload_json, candidate_id FROM candidates WHERE {where}",  # noqa: S608 - static clause
        parameters,
    ).fetchall()
    remaining_releases: set[str] = set()
    for row in remaining_rows:
        payload = json.loads(row["payload_json"])
        track = payload.get("track") if isinstance(payload.get("track"), dict) else {}
        audio = payload.get("audio") if isinstance(payload.get("audio"), dict) else {}
        remaining_releases.add(
            normalize_upc(track.get("upc"))
            or clean(audio.get("folder_id"))
            or normalize_text(track.get("release"))
            or row["candidate_id"]
        )
    counts["remaining"] = len(remaining_rows)
    counts["remainingReleases"] = len(remaining_releases)
    return counts


def export_private_manifests(connection: sqlite3.Connection, output_dir: Path) -> dict[str, int]:
    counts = {"exact": 0, "review": 0, "quarantine": 0, "ownerDirect": 0}
    handles = {
        status: (output_dir / f"{status}.jsonl").open("w", encoding="utf-8", newline="\n")
        for status in ("exact", "review", "quarantine")
    }
    direct = (output_dir / "catalog-owner-direct.jsonl").open("w", encoding="utf-8", newline="\n")
    manifest = (output_dir / "manifest.jsonl").open("w", encoding="utf-8", newline="\n")
    try:
        for row in connection.execute("SELECT * FROM candidates ORDER BY candidate_id"):
            payload = json.loads(row["payload_json"])
            record = {
                **payload,
                "status": row["status"],
                "reasons": json.loads(row["reasons_json"]),
                "inspection": {
                    "status": row["inspection_status"],
                    "mode": row["inspection_mode"],
                    "content_type": row["content_type"],
                    "content_length": row["content_length"],
                    "wav": json.loads(row["wav_json"]) if row["wav_json"] else None,
                    "sha256": row["sha256"],
                    "error": row["error"],
                },
            }
            line = json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n"
            manifest.write(line)
            handles[row["status"]].write(line)
            counts[row["status"]] += 1
            if direct_publication_eligible(record):
                direct.write(line)
                counts["ownerDirect"] += 1
    finally:
        manifest.close()
        direct.close()
        for handle in handles.values():
            handle.close()
    (output_dir / "summary.json").write_text(
        json.dumps({"generated_at": dt.datetime.now(dt.timezone.utc).isoformat(), "counts": counts}, indent=2) + "\n",
        encoding="utf-8",
    )
    return counts


def merge_drive_discovery(
    release_audio: Sequence[ReleaseAudio],
    inventory_by_release: Mapping[str, list[DriveAudio]],
    inventory_children: Mapping[str, list[dict[str, Any]]],
    drive: GoogleDriveClient,
    *,
    discover: bool,
    inventory_complete: bool = False,
) -> list[ReleaseAudio]:
    """Merge the workbook references with the newest Drive snapshot.

    Drive inventory rows are authoritative for a matching file ID.  In
    particular, ``size`` and ``modifiedTime`` must replace the older workbook
    values because they are part of the candidate fingerprint that decides
    whether a completed checksum can be reused.

    A partial inventory is additive and can never imply a deletion.  Only an
    explicitly complete inventory replaces the workbook's file list for a
    release folder; this makes removals fail closed without treating an
    interrupted crawl as a catalogue deletion.
    """

    merged: list[ReleaseAudio] = []
    for release in release_audio:
        inventory_files = inventory_by_release.get(release.folder_id, [])
        if inventory_complete and release.folder_id:
            files: dict[str, DriveAudio] = {}
        else:
            files = {audio.file_id: audio for audio in release.files}
        for audio in inventory_files:
            # Assignment is deliberate: the crawler is fresher than workbook
            # hyperlinks, even when Google keeps the same immutable file ID.
            files[audio.file_id] = audio
        if discover and release.folder_id and not inventory_complete:
            for audio in drive.discover_wavs(release.folder_id, inventory_children=inventory_children):
                files[audio.file_id] = audio
        merged.append(dataclasses.replace(release, files=tuple(files.values())))
    return merged


def filter_release_scope(
    tracks: Sequence[Track], release_audio: Sequence[ReleaseAudio], release_filters: Sequence[str]
) -> tuple[list[Track], list[ReleaseAudio]]:
    if not release_filters:
        return list(tracks), list(release_audio)
    wanted = {normalize_text(value) for value in release_filters}
    return (
        [track for track in tracks if track.normalized_release in wanted],
        [release for release in release_audio if normalize_text(release.release) in wanted],
    )


def select_smoke_candidates(candidates: Sequence[Candidate]) -> list[Candidate]:
    selected: list[Candidate] = []
    for release in ("time", "rise", "signal flow"):
        candidates_for_release = [
            candidate
            for candidate in candidates
            if candidate.track and candidate.audio and normalize_text(candidate.track.release) == release
        ]
        if candidates_for_release:
            selected.append(sorted(candidates_for_release, key=lambda candidate: (-float(candidate.audio_match_score or 0), candidate.candidate_id))[0])
    return selected


def assert_private_output(output_dir: Path) -> None:
    repository = Path(__file__).resolve().parents[1]
    resolved = output_dir.resolve()
    try:
        relative = resolved.relative_to(repository)
    except ValueError:
        return
    if not relative.parts[:2] == ("catalog-audit", "private"):
        raise ValueError("Private manifests inside the repository must stay under catalog-audit/private/.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--workbook", type=Path, required=True, help="Private All DATA.xlsx export.")
    parser.add_argument("--orchard", type=Path, help="Private The Orchard Spotify mapping export (XLSX/CSV/TSV/JSON).")
    parser.add_argument("--drive-inventory", type=Path, help="Optional private connector export for recursive Drive discovery.")
    parser.add_argument(
        "--central-drive-inventory",
        type=Path,
        help="Optional additive private connector snapshot of the flat central Fichiers/Artwork folders.",
    )
    parser.add_argument(
        "--central-drive-baseline-inventory",
        type=Path,
        help="Optional prior central snapshot whose established associations must remain immutable.",
    )
    parser.add_argument("--output-dir", type=Path, default=PRIVATE_DIRECTORY, help="Ignored private working directory.")
    parser.add_argument("--release", action="append", default=[], help="Limit to a release name; repeat for several releases.")
    parser.add_argument("--smoke", action="store_true", help="Select one WAV each from Time, Rise and Signal Flow.")
    parser.add_argument("--discover-drive", action="store_true", help="Recursively list Drive folders missing from workbook audio links.")
    parser.add_argument("--inspect", choices=("none", "range", "full"), default="none", help="WAV verification depth.")
    parser.add_argument("--batch-size", type=int, default=25, help="Maximum audio files inspected in this invocation.")
    parser.add_argument(
        "--release-batch-size",
        type=int,
        help="Maximum distinct releases inspected in this invocation; omitted means no release cap.",
    )
    parser.add_argument(
        "--max-inspection-seconds",
        type=int,
        help="Soft wall-clock budget checked between files; a current Drive read is allowed to finish.",
    )
    parser.add_argument("--allow-network", action="store_true", help="Explicitly allow Drive reads for inspection/discovery.")
    parser.add_argument("--force", action="store_true", help="Reinspect candidates even when the fingerprint is unchanged.")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--apply", action="store_true", help="Write/resume the ignored SQLite state and private manifests.")
    mode.add_argument("--dry-run", action="store_true", help="Only print aggregate plan counts (the safe default).")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    arguments = build_parser().parse_args(argv)
    apply = bool(arguments.apply)
    if arguments.batch_size < 1:
        raise ValueError("--batch-size must be at least 1")
    if arguments.release_batch_size is not None and arguments.release_batch_size < 1:
        raise ValueError("--release-batch-size must be at least 1")
    if arguments.max_inspection_seconds is not None and arguments.max_inspection_seconds < 60:
        raise ValueError("--max-inspection-seconds must be at least 60")
    if (arguments.inspect != "none" or arguments.discover_drive) and not arguments.allow_network:
        raise ValueError("Network operations require the explicit --allow-network flag.")
    if not apply and arguments.inspect != "none":
        raise ValueError("Inspection writes resumable state; combine --inspect with --apply.")
    if arguments.smoke:
        arguments.release = ["Time", "Rise", "Signal Flow"]
        arguments.batch_size = min(arguments.batch_size, 3)

    output_dir = arguments.output_dir
    assert_private_output(output_dir)
    tracks, release_audio, covers = load_workbook_sources(arguments.workbook)
    orchard = merge_orchard_rows(load_orchard(arguments.orchard), load_identity_info(arguments.workbook))
    inventory_by_release, inventory_children = load_drive_inventory(arguments.drive_inventory)
    central_audio, central_artwork, central_complete = load_central_drive_inventory(
        arguments.central_drive_inventory
    )
    baseline_central_audio, _baseline_artwork, _baseline_complete = load_central_drive_inventory(
        arguments.central_drive_baseline_inventory or arguments.central_drive_inventory
    )
    drive = GoogleDriveClient(clean(os.environ.get("GOOGLE_DRIVE_ACCESS_TOKEN")))
    release_audio = merge_drive_discovery(
        release_audio,
        inventory_by_release,
        inventory_children,
        drive,
        discover=bool(arguments.discover_drive),
        inventory_complete=drive_inventory_is_complete(arguments.drive_inventory),
    )
    covers = merge_drive_cover_fallbacks(covers, release_audio, inventory_children)
    baseline_central_wav = [
        audio
        for audio in baseline_central_audio
        if (audio_source_descriptor(audio) or (None, None))[0] == "wav"
    ]
    release_audio, baseline_audio_summary = merge_central_audio_mappings(
        tracks, release_audio, baseline_central_wav
    )
    covers, central_cover_summary = merge_central_cover_mappings(
        covers, tracks, central_artwork
    )
    tracks, release_audio = filter_release_scope(tracks, release_audio, arguments.release)
    candidates = build_candidates(tracks, release_audio, covers, orchard)
    baseline_direct_rows = {
        candidate.track.source_row
        for candidate in candidates
        if candidate.track is not None
        and direct_publication_eligible(
            {
                **candidate_payload(candidate),
                "status": candidate.status,
                "reasons": candidate.reasons,
            }
        )
    }
    baseline_file_ids = {
        audio.file_id for group in release_audio for audio in group.files
    }.union(audio.file_id for audio in baseline_central_wav)
    central_audio_pins, pin_summary = build_central_audio_pins(
        tracks,
        candidates,
        central_audio,
        excluded_file_ids=baseline_file_ids,
    )
    candidates = apply_central_audio_pins(candidates, central_audio_pins)
    final_direct_rows = {
        candidate.track.source_row
        for candidate in candidates
        if candidate.track is not None
        and direct_publication_eligible(
            {
                **candidate_payload(candidate),
                "status": candidate.status,
                "reasons": candidate.reasons,
            }
        )
    }
    if not baseline_direct_rows.issubset(final_direct_rows):
        raise RuntimeError("central_audio_pin_non_monotonic")
    central_audio_summary = {
        "available": len(central_audio),
        "mapped": baseline_audio_summary["mapped"] + pin_summary["pinned"],
        "baselineMapped": baseline_audio_summary["mapped"],
        "baselineOwnerDirect": len(baseline_direct_rows),
        "ownerDirectAdded": len(final_direct_rows.difference(baseline_direct_rows)),
        **{key: value for key, value in pin_summary.items() if key != "available"},
    }
    if arguments.smoke:
        candidates = select_smoke_candidates(candidates)

    state_candidates = list(
        {candidate.candidate_id: candidate for candidate in candidates}.values()
    )
    planned_counts = {
        status: sum(candidate.status == status for candidate in state_candidates)
        for status in ("exact", "review", "quarantine")
    }
    owner_direct_candidates = sum(
        direct_publication_eligible(
            {
                **candidate_payload(candidate),
                "status": candidate.status,
                "reasons": candidate.reasons,
            }
        )
        for candidate in state_candidates
    )
    print(
        json.dumps(
            {
                "mode": "apply" if apply else "dry-run",
                "tracks": len(tracks),
                "track_candidates": sum(
                    candidate.track is not None for candidate in state_candidates
                ),
                "release_audio_groups": len(release_audio),
                "candidates": len(state_candidates),
                "planned_status": planned_counts,
                "owner_direct_candidates": owner_direct_candidates,
                "orchard_rows": len(orchard),
                "covers": len(covers),
                "central_inventory": {
                    "complete": central_complete,
                    "audio": central_audio_summary,
                    "artwork": central_cover_summary,
                },
            },
            indent=2,
        )
    )
    if not apply:
        return 0

    output_dir.mkdir(parents=True, exist_ok=True)
    connection = open_state(output_dir / "ingestion-state.sqlite3")
    try:
        upsert_candidates(connection, candidates, force=bool(arguments.force))
        stale_quarantined = 0
        if (
            source_inventories_are_complete(
                arguments.drive_inventory,
                arguments.central_drive_inventory,
                central_complete,
            )
            and not arguments.release
            and not arguments.smoke
        ):
            stale_quarantined = quarantine_stale_candidates(
                connection, {candidate.candidate_id for candidate in candidates}
            )
        inspected = inspect_pending(
            connection,
            drive,
            mode=arguments.inspect,
            batch_size=arguments.batch_size,
            force=bool(arguments.force),
            release_batch_size=arguments.release_batch_size,
            maximum_seconds=arguments.max_inspection_seconds,
        )
        inspected["staleQuarantined"] = stale_quarantined
        final_counts = export_private_manifests(connection, output_dir)
    finally:
        connection.close()
    print(json.dumps({"inspection": inspected, "manifest": final_counts}, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted. Completed files are checkpointed and the next run will resume.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as error:  # concise CLI boundary; details remain out of public manifests
        print(f"catalog ingestion failed: {type(error).__name__}: {error}", file=sys.stderr)
        raise SystemExit(1)
